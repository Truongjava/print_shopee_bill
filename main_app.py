"""
ShopeePrint — Desktop App Shopee Seller Automation + Bill Calculate
====================================================================
GUI: PySide6 (Qt for Python) — 4-tab layout with QSS stylesheet (SaaS Light Theme)
Chạy: python main_app.py
Đóng gói .exe: pyinstaller TTS_Bill.spec
"""
import os, sys, json, shutil, threading, re as _re
from datetime import datetime, timedelta
from pathlib import Path

# ═══════════════════════════════════════════════════════════
# PySide6 imports
# ═══════════════════════════════════════════════════════════
from PySide6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
    QLabel, QLineEdit, QPushButton, QGroupBox, QCheckBox, QRadioButton,
    QButtonGroup, QSpinBox, QComboBox, QTextEdit, QListWidget,
    QListWidgetItem, QScrollArea, QStackedWidget, QTableWidget,
    QTableWidgetItem, QHeaderView, QFileDialog, QMessageBox,
    QFrame, QApplication,
)
from PySide6.QtCore import (
    Qt, Signal, Slot, QThread, QTimer, QObject, QMetaObject,
)
from PySide6.QtGui import (
    QFont, QTextCursor,
)

# ═══════════════════════════════════════════════════════════
# Lazy imports: playwright & calculator chỉ import khi cần dùng
# để tránh khởi động app chậm (mỗi cái ~2-4 giây)
# - sync_playwright → import trong run_automation()
# - calculator modules → import trong run_calculator()

# ═══════════════════════════════════════════════════════════
# Paths & config
# ═══════════════════════════════════════════════════════════
BASE_DIR = Path(__file__).parent
BILL_DIR = BASE_DIR / 'bill_calculate'
UPLOAD_DIR = BILL_DIR / 'uploads'
sys.path.insert(0, str(BILL_DIR))

# Lazy import: calculator modules sẽ được import trong run_calculator()
# để tránh kéo pdfplumber + openpyxl + win32com (~6s) lúc khởi động

_calculator_cache = None

def _get_calculator():
    """Lazy import calculator module — chỉ load khi cần, cache lại cho lần sau."""
    global _calculator_cache
    if _calculator_cache is None:
        from calculator import process_all, extract_report_data, aggregate_reports, detect_pdf_type, split_shopee_pdf
        _calculator_cache = (process_all, extract_report_data, aggregate_reports, detect_pdf_type, split_shopee_pdf)
    return _calculator_cache

# ═══════════════════════════════════════════════════════════
# Frozen / source mode — detect paths
# ═══════════════════════════════════════════════════════════
if getattr(sys, 'frozen', False):
    BASE_DIR = Path(sys.executable).parent
    BILL_DIR = BASE_DIR / 'bill_calculate'
    UPLOAD_DIR = BILL_DIR / 'uploads'
    DEFAULT_COOKIE = Path(sys._MEIPASS) / 'banhang.shopee.vn_24-07-2026.json'
    MASTER_DEFAULT = Path(sys._MEIPASS) / 'mã combo.xlsx'
    RETAIL_DEFAULT = Path(sys._MEIPASS) / 'sp bán lẻ.xlsx'
    TEMPLATE_DEFAULT = Path(sys._MEIPASS) / 'Bảng thống kê hàng.xlsx'
else:
    DEFAULT_COOKIE = BASE_DIR / 'banhang.shopee.vn_24-07-2026.json'
    MASTER_DEFAULT = BASE_DIR / 'mã combo.xlsx'
    RETAIL_DEFAULT = BASE_DIR / 'sp bán lẻ.xlsx'
    TEMPLATE_DEFAULT = BASE_DIR / 'Bảng thống kê hàng.xlsx'

TARGET_URL = 'https://banhang.shopee.vn'
ORDERS_URL = 'https://banhang.shopee.vn/portal/sale/mass/ship?mass_shipment_tab=201&filter.order_item_filter_type=item0&filter.order_process_status=1&filter.sort.sort_type=2&filter.sort.ascending=true&filter.pre_order=2&filter.shipping_priority=2&filter.entity_type=1'
# Shopee shipping methods
CARRIER_URLS = {
    'SPX Express':                ORDERS_URL + '&filter.shipping_method=50021',
    'Giao Hàng Nhanh':            ORDERS_URL + '&filter.shipping_method=14',
    'Ninja Van':                  ORDERS_URL + '&filter.shipping_method=50023',
    'VNPost Nhanh':               ORDERS_URL + '&filter.shipping_method=29',
    'J&T Express':                ORDERS_URL + '&filter.shipping_method=50018',
    'VNP - Hàng Cồng Kềnh':       ORDERS_URL + '&filter.shipping_method=30',
    'Ahamove SBS - Trong Ngày':   ORDERS_URL + '&filter.shipping_method=50033',
    'GHN - Hàng Cồng Kềnh':       ORDERS_URL + '&filter.shipping_method=50032',
    'Ahamove - Trong Ngày':       ORDERS_URL + '&filter.shipping_method=50044',
    'BEST Express':               ORDERS_URL + '&filter.shipping_method=50024',
    'NJV - Hàng Cồng Kềnh':       ORDERS_URL + '&filter.shipping_method=50034',
    'SPX - Hàng Cồng Kềnh':       ORDERS_URL + '&filter.shipping_method=50025',
    'SPX Express - Trong Ngày':   ORDERS_URL + '&filter.shipping_method=50041',
    'SPX Express SBS - Trong Ngày': ORDERS_URL + '&filter.shipping_method=50051',
    'Tủ nhận hàng - SPX':         ORDERS_URL + '&filter.shipping_method=50039',
    'VTP - Hàng Cồng Kềnh':       ORDERS_URL + '&filter.shipping_method=50100',
    'Viettel Post':               ORDERS_URL + '&filter.shipping_method=50037',
    'Đơn vị vận chuyển khác':     ORDERS_URL + '&filter.shipping_method=0',
}
BATCH_SIZE = 50
os.makedirs(UPLOAD_DIR, exist_ok=True)


# ============================================================
# AUTOMATION
# ============================================================
def _detect_captcha(page, log_cb, state_cb, stop_event):
    """Kiểm tra xem Shopee có hiện CAPTCHA không. Nếu có → dừng chờ user giải."""
    captcha_selectors = [
        # Shopee slider / image CAPTCHA
        'iframe[src*="captcha"]',
        'iframe[src*="verify"]',
        'div[class*="captcha"]',
        'div[class*="verify"]',
        'div[class*="slider"]',
        '.shopee-captcha',
        '#captcha',
        '#captcha-container',
        '.captcha_verify',
        # Shopee specific
        '[class*="sec-captcha"]',
        '[class*="security-check"]',
        '[class*="verification"]',
        # Text-based detection
        'text=Kéo thanh trượt',
        'text=Kéo để xác',
        'text=Trượt để xác',
        'text=Slide to verify',
        'text=Please verify',
        'text=Verify you are human',
        'text=Xác minh',
        'text=Xác minh bạn không phải',
        'text=Vui lòng xác minh',
        'text=Security verification',
    ]
    for selector in captcha_selectors:
        try:
            el = page.locator(selector).first
            if el.count() > 0 and el.is_visible(timeout=1000):
                # Có CAPTCHA!
                log_cb('🛑 PHÁT HIỆN CAPTCHA! Vui lòng kéo hình xác minh trên Chrome...', 'err')
                state_cb('captcha', '⏳ Đợi bạn giải CAPTCHA...')
                # Chụp màn hình
                try:
                    ss = f'captcha_{datetime.now().strftime("%m-%d_%H-%M-%S")}.png'
                    page.screenshot(path=ss)
                    log_cb(f'  📸 Screenshot: {ss}', 'info')
                except: pass
                # Đợi user giải CAPTCHA (polling mỗi 2s, tối đa 5 phút)
                import time as _t
                for _ in range(150):  # 150 × 2s = 5 phút
                    if stop_event and stop_event.is_set():
                        return
                    _t.sleep(2)
                    # Kiểm tra CAPTCHA đã biến mất chưa
                    try:
                        if el.count() == 0 or not el.is_visible(timeout=500):
                            log_cb('✅ CAPTCHA đã được giải — tiếp tục...', 'ok')
                            state_cb('running', 'Đang tiếp tục...')
                            return
                    except:
                        log_cb('✅ CAPTCHA đã được giải — tiếp tục...', 'ok')
                        state_cb('running', 'Đang tiếp tục...')
                        return
                log_cb('⚠ Hết thời gian chờ CAPTCHA — thử tiếp...', 'warn')
                return
        except Exception:
            continue

def run_automation(cookie_path, output_dir, max_orders, log_cb, state_cb, stop_event=None,
                   existing_playwright=None, existing_browser=None, carrier=None, test_mode=False,
                   exclude_pre_orders=True):
    """
    Shopee Automation: đăng nhập → chọn carrier → chọn tất cả đơn → chuẩn bị hàng → tải PDF.
    """
    with open(cookie_path, 'r', encoding='utf-8') as f:
        cd = json.load(f)
    cookies_list = cd.get('cookies', cd if isinstance(cd, list) else [])
    pdf_files = []
    carrier_counts: dict[str, int] = {}  # khởi tạo sớm, tránh NameError ở except
    carrier_label = f' [{carrier}]' if carrier else ''
    orders_url = CARRIER_URLS.get(carrier, ORDERS_URL)
    if stop_event is None:
        stop_event = threading.Event()

    # ── Khởi tạo / tái sử dụng browser ──
    browser_ok = False
    if existing_browser and existing_playwright:
        try:
            playwright = existing_playwright
            browser = existing_browser
            if not browser.is_connected():
                log_cb('⚠ Browser cũ đã ngắt kết nối — tạo mới...', 'warn')
                raise RuntimeError('browser disconnected')
            context = browser.contexts[0] if browser.contexts else browser.new_context(
                viewport={'width': 1366, 'height': 768},
                user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/131.0.0.0 Safari/537.36',
                accept_downloads=True)
            page = context.new_page()
            log_cb('♻ Dùng lại browser — tạo tab mới...', 'info')
            for p in list(context.pages):
                if p != page:
                    try:
                        if not p.is_closed():
                            p.close()
                    except Exception:
                        pass
            page.goto(orders_url, wait_until='domcontentloaded', timeout=60000)
            page.wait_for_timeout(4000)
            _detect_captcha(page, log_cb, state_cb, stop_event)
            browser_ok = True
        except Exception as e:
            log_cb(f'⚠ Không dùng lại được browser cũ ({e}) — tạo mới...', 'warn')
            try:
                existing_browser.close()
            except Exception:
                pass
            import time as _time
            _time.sleep(0.5)
            existing_browser = None
            existing_playwright = None

    if not browser_ok:
        try:
            from playwright.sync_api import sync_playwright  # lazy import (~2.3s)
            playwright = sync_playwright().start()
            launch_opts = {'headless': False, 'channel': 'chrome',
                           'args': ['--disable-blink-features=AutomationControlled']}
            log_cb('🌐 Dùng Google Chrome có sẵn trên máy', 'info')
            browser = playwright.chromium.launch(**launch_opts)
        except Exception as e:
            log_cb(f'✗ Không thể khởi động browser: {e}', 'err')
            raise RuntimeError(f'Không thể khởi động Chromium: {e}') from e
        context = browser.new_context(
            viewport={'width': 1366, 'height': 768},
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/131.0.0.0 Safari/537.36',
            accept_downloads=True)
        pw_cookies = []
        for c in cookies_list:
            if not c.get('name') or not c.get('value'): continue
            pw = {'name': c['name'], 'value': c['value'],
                  'domain': c.get('domain', '.shopee.vn'), 'path': c.get('path', '/')}
            if c.get('expirationDate') and not c.get('session'):
                pw['expires'] = int(c['expirationDate'])
            if 'httpOnly' in c: pw['httpOnly'] = c['httpOnly']
            if 'secure' in c: pw['secure'] = c['secure']
            if c.get('sameSite'):
                pw['sameSite'] = {'strict': 'Strict', 'lax': 'Lax',
                                  'no_restriction': 'None', 'unspecified': 'Lax'}.get(c['sameSite'], 'Lax')
            pw_cookies.append(pw)
        context.add_cookies(pw_cookies)
        page = context.new_page()
        page.goto(TARGET_URL, wait_until='domcontentloaded', timeout=30000)
        page.wait_for_timeout(2000)
        _detect_captcha(page, log_cb, state_cb, stop_event)

    try:
        # ── Navigate đến trang đơn hàng của carrier ──
        log_cb(f'📥 Tải đơn{carrier_label}...', 'info')
        state_cb('navigating', f'Đang tải danh sách đơn{carrier_label}...')
        for goto_attempt in range(3):
            try:
                page.goto(orders_url, wait_until='domcontentloaded', timeout=60000)
                break
            except Exception as e:
                if goto_attempt < 2:
                    log_cb(f'  ⚠ Lỗi mạng (lần {goto_attempt + 1}/3): {e} — thử lại sau 5s...', 'warn')
                    page.wait_for_timeout(5000)
                else:
                    raise
        # Đợi React/Vue render xong danh sách đơn hàng (polling)
        for _ in range(15):  # tối đa 15 giây
            cb_count = page.evaluate(
                "() => document.querySelectorAll('input[type=\"checkbox\"]').length")
            if cb_count > 1:  # Có ít nhất select-all + 1 đơn hàng
                break
            page.wait_for_timeout(1000)
        page.wait_for_timeout(2000)
        _detect_captcha(page, log_cb, state_cb, stop_event)

        # ── 0. Quét số đơn từng carrier từ filter panel (1 lần) ──
        carrier_counts = {}
        try:
            filter_text = page.locator('.order-container').first.inner_text()
            import re as _re2
            for line in filter_text.split('\n'):
                m = _re2.match(r'^(.+?)\s*\((\d+)\)$', line.strip())
                if m and m.group(1) in CARRIER_URLS:
                    carrier_counts[m.group(1)] = int(m.group(2))
            if carrier_counts:
                log_cb(f'  📊 Carrier có đơn: { {k: v for k, v in carrier_counts.items() if v > 0} }', 'info')
        except Exception:
            pass

        # ════════════════════════════════════════════════════════════════
        # PAGINATION LOOP - xử lý từng batch 200 đơn đến khi hết
        # ════════════════════════════════════════════════════════════════
        all_batch_pdfs = []
        PAGE_SIZE = 200

        # Tính số batch dự kiến từ scan ban đầu (carrier_counts)
        if carrier_counts and carrier in carrier_counts:
            total_orders = carrier_counts[carrier]
            expected_batches = (total_orders + PAGE_SIZE - 1) // PAGE_SIZE
            max_batches = expected_batches
            log_cb(f'  📊 {carrier}: {total_orders} đơn → {expected_batches} batch', 'dim')
        else:
            max_batches = 50  # fallback nếu không scan được

        for batch_num in range(1, max_batches + 1):
            # ── Reload orders page (batch 1 already loaded above) ──
            if batch_num > 1:
                log_cb(f'📦 Batch {batch_num}: tải lại danh sách đơn{carrier_label}...', 'info')
                for goto_attempt in range(3):
                    try:
                        page.goto(orders_url, wait_until='domcontentloaded', timeout=60000)
                        break
                    except Exception as e:
                        if goto_attempt < 2:
                            log_cb(f'  ⚠ Lỗi mạng (lần {goto_attempt + 1}/3): {e} — thử lại sau 5s...', 'warn')
                            page.wait_for_timeout(5000)
                        else:
                            raise
                for _ in range(15):
                    if page.evaluate("() => document.querySelectorAll('input[type=\"checkbox\"]').length") > 1:
                        break
                    page.wait_for_timeout(1000)
                page.wait_for_timeout(2000)
                _detect_captcha(page, log_cb, state_cb, stop_event)
            else:
                log_cb(f'📦 Batch 1: xử lý đơn{carrier_label}...', 'info')

            # ── 1. Dismiss popup ──
            page.wait_for_timeout(3000)
            try:
                dismissed = page.evaluate('''() => {
                    const buttons = document.querySelectorAll('button');
                    for (const btn of buttons) {
                        const txt = btn.textContent.trim();
                        if (txt === 'Đã hiểu' || txt === 'Bỏ qua' || txt === 'Skip') {
                            btn.click();
                            btn.dispatchEvent(new MouseEvent('click', {bubbles: true}));
                            return txt;
                        }
                    }
                    return '';
                }''')
                if dismissed:
                    log_cb(f'  ✓ Đã tắt popup "{dismissed}"', 'info')
                    page.wait_for_timeout(2000)
            except Exception:
                pass

            # ── 2. Select 200/page ──
            try:
                current_size = page.evaluate(
                    '() => (document.querySelector("button.pagination-sizes__content") || {}).textContent || ""'
                ).strip()
                if current_size and not current_size.startswith('200'):
                    page.evaluate('''() => {
                        const btn = document.querySelector('button.pagination-sizes__content');
                        if (btn) {
                            ['mousedown', 'mouseup', 'click'].forEach(t => {
                                btn.dispatchEvent(new MouseEvent(t, {bubbles: true, cancelable: true, view: window}));
                            });
                        }
                    }''')
                    page.wait_for_timeout(1500)
                    page.evaluate('''() => {
                        const items = document.querySelectorAll('li.page-size-option');
                        for (const item of items) {
                            if (item.textContent.trim() === '200') { item.click(); return; }
                        }
                    }''')
                    page.wait_for_timeout(5000)
                    log_cb('  ✓ Đã chọn 200 đơn/trang', 'info')
            except Exception as e:
                log_cb(f'  ⚠ Không chọn được 200/trang: {e}', 'dim')

            # ── 3. Check remaining orders ──
            remaining = page.evaluate(
                "() => document.querySelectorAll('input[type=\"checkbox\"]').length") - 1
            if remaining <= 0:
                log_cb(f'  ✓ Hết đơn khả dụng{carrier_label} — hoàn thành!', 'ok')
                break

            log_cb(f'  📋 Batch {batch_num}: {remaining} đơn còn lại', 'info')

            # ── 4. Click Select All ──
            page.evaluate('''() => {
                const label = document.querySelector('[data-testid="mass-ship-checkbox-all"]');
                if (!label) return;
                const indicator = label.querySelector('.eds-checkbox__indicator');
                const target = indicator || label;
                ['mousedown', 'mouseup', 'click'].forEach(t => {
                    target.dispatchEvent(new MouseEvent(t, {bubbles: true, cancelable: true, view: window}));
                });
                const input = label.querySelector('input[type="checkbox"]');
                if (input) input.dispatchEvent(new Event('change', {bubbles: true}));
            }''')
            page.wait_for_timeout(5000)

            checked = page.evaluate(
                "() => { const all = document.querySelectorAll('input[type=\"checkbox\"]:checked'); let c = 0; for (const cb of all) { if (!cb.closest('[data-testid=\"mass-ship-checkbox-all\"]')) c++; } return c; }")
            log_cb(f'  ✓ Đã chọn {checked}/{remaining} đơn hàng', 'ok')

            if checked == 0:
                log_cb('  ✓ Hết đơn — hoàn thành.', 'ok')
                break

            # ── 5. Test mode ──
            if test_mode:
                log_cb(f'  🧪 TEST MODE: Dừng tại bước chọn đơn — không in.', 'warn')
                try:
                    ss = str(Path(output_dir) / f'test_mode_{carrier or "all"}_batch{batch_num}.png')
                    page.screenshot(path=ss)
                    log_cb(f'  📸 Screenshot: {ss}', 'info')
                except Exception:
                    pass
                break

            # ── 5.5. Delay trước khi click pickup ──
            page.wait_for_timeout(5000)

            # ── 6. Click pickup button ──
            state_cb('printing', 'Đang yêu cầu lấy hàng...')
            pickup_btn = None
            for btn_text in ['Yêu cầu đơn vị vận chuyển đến lấy hàng',
                             'Yêu cầu đơn vị vận chuyển', 'Yêu cầu lấy hàng']:
                try:
                    btn = page.locator(f'button:has-text("{btn_text}")').first
                    if btn.count() > 0 and btn.is_visible(timeout=2000):
                        pickup_btn = btn
                        break
                except Exception:
                    pass
            if not pickup_btn:
                try:
                    btn = page.locator('[data-testid="arrange-pickup-confirm-button"]').first
                    if btn.count() > 0 and btn.is_visible(timeout=2000):
                        pickup_btn = btn
                except Exception:
                    pass
            if pickup_btn:
                pickup_btn.click(force=True, timeout=5000)
                log_cb('  ✓ Đã bấm "Yêu cầu đơn vị vận chuyển đến lấy hàng"', 'ok')
                log_cb('  ⏳ Đợi 30 giây cho popup hiện ra...', 'info')
                page.wait_for_timeout(30000)
            else:
                log_cb('  ✗ Không tìm thấy nút "Yêu cầu lấy hàng"', 'err')
                page.screenshot(path=str(Path(output_dir) / f'debug_no_pickup_btn_{carrier or "all"}_batch{batch_num}.png'))
                continue

            # ── 7. Click "Tạo" in popup ──
            state_cb('printing', 'Đang tạo yêu cầu...')
            tao_btn = None
            for _ in range(30):
                for btn_text in ['Tạo', 'Xác nhận', 'Confirm', 'Create']:
                    try:
                        btn = page.locator(f'button:has-text("{btn_text}")').first
                        if btn.count() > 0 and btn.is_visible(timeout=500):
                            tao_btn = btn
                            break
                    except Exception:
                        pass
                if tao_btn:
                    break
                page.wait_for_timeout(1000)
            if tao_btn:
                tao_btn.click(force=True, timeout=5000)
                log_cb('  ✓ Đã bấm "Tạo"', 'ok')
                page.wait_for_timeout(5000)
            else:
                log_cb('  ⚠ Không tìm thấy nút "Tạo" — thử batch tiếp...', 'warn')
                continue

            # ── 8. Click manual select ──
            state_cb('printing', 'Đang chọn loại phiếu...')
            log_cb('  🔍 Đang tìm nút "Chọn phiếu thủ công"...', 'dim')
            manual_btn_found = False
            for _ in range(30):
                try:
                    clicked = page.evaluate('''() => {
                        const btns = document.querySelectorAll('button');
                        for (const btn of btns) {
                            if (btn.textContent.includes('Chọn phiếu thủ công') && btn.offsetParent !== null) {
                                btn.click();
                                return true;
                            }
                        }
                        return false;
                    }''')
                    if clicked:
                        log_cb('  ✓ Đã bấm "Chọn phiếu thủ công"', 'ok')
                        manual_btn_found = True
                        page.wait_for_timeout(2000)
                        break
                except Exception:
                    pass
                page.wait_for_timeout(1000)
            if not manual_btn_found:
                log_cb('  ⚠ Không tìm thấy nút "Chọn phiếu thủ công" sau 30 giây', 'warn')
                try:
                    page.screenshot(path=str(Path(output_dir) / f'debug_no_manual_btn_{carrier or "all"}_batch{batch_num}.png'))
                except Exception:
                    pass
                continue

            # ── 9. Tick checkboxes ──
            state_cb('printing', 'Đang tích chọn loại phiếu...')
            CHECK_TEXTS = ['Phiếu xuất hàng', 'Phiếu gửi hàng và phiếu đóng gói']

            # 9a. Dump + un-tick
            try:
                all_checkboxes_info = page.evaluate('''() => {
                    const results = [];
                    const labels = document.querySelectorAll('label');
                    for (const lbl of labels) {
                        const cb = lbl.querySelector('input[type="checkbox"]');
                        if (cb && lbl.offsetParent !== null) {
                            results.push({text: lbl.textContent.trim().substring(0, 120), checked: cb.checked});
                        }
                    }
                    let unTicked = 0;
                    for (const lbl of labels) {
                        const cb = lbl.querySelector('input[type="checkbox"]');
                        if (cb && cb.checked && cb.offsetParent !== null) {
                            lbl.click(); cb.dispatchEvent(new Event('change', {bubbles: true})); unTicked++;
                        }
                    }
                    return JSON.stringify({checkboxes: results, unTicked: unTicked});
                }''')
                info = json.loads(all_checkboxes_info)
                for cbi in info.get('checkboxes', []):
                    text = cbi['text'].strip()
                    if not text:
                        continue  # bỏ qua checkbox rỗng (ẩn, không có text)
                    marker = '☑' if cbi['checked'] else '☐'
                    log_cb(f'  {marker} {text}', 'dim')
                if info.get('unTicked', 0) > 0:
                    log_cb(f'  🗑 Đã bỏ tick {info["unTicked"]} checkbox', 'dim')
                page.wait_for_timeout(500)
            except Exception as e:
                log_cb(f'  ⚠ Lỗi un-tick: {e}', 'warn')

            # 9b. Tick desired checkboxes
            ticked_texts = []
            for wanted in CHECK_TEXTS:
                wanted_norm = wanted.strip().lower()
                try:
                    ticked = page.evaluate(f'''(wanted) => {{
                        const norm = (s) => s.trim().toLowerCase().replace(/\\s+/g, ' ');
                        const labels = document.querySelectorAll('label');
                        for (const lbl of labels) {{
                            const cb = lbl.querySelector('input[type="checkbox"]');
                            if (!cb || lbl.offsetParent === null) continue;
                            if (norm(lbl.textContent).includes(wanted)) {{
                                if (!cb.checked) {{ lbl.click(); cb.dispatchEvent(new Event('change', {{bubbles: true}})); }}
                                return norm(lbl.textContent).substring(0, 80) + ' => ' + (cb.checked ? 'OK' : 'FAIL');
                            }}
                        }}
                        const walker = document.createTreeWalker(document.body, NodeFilter.SHOW_TEXT); let node;
                        while (node = walker.nextNode()) {{
                            if (norm(node.textContent).includes(wanted) && node.parentElement.offsetParent !== null) {{
                                let el = node.parentElement;
                                for (let i = 0; i < 5 && el; i++) {{
                                    const cb = el.querySelector('input[type="checkbox"]');
                                    if (cb) {{
                                        if (!cb.checked) {{ el.click(); cb.dispatchEvent(new Event('change', {{bubbles: true}})); }}
                                        return 'treeWalker=>' + wanted + ' => ' + (cb.checked ? 'OK' : 'FAIL');
                                    }}
                                    el = el.parentElement;
                                }}
                            }}
                        }}
                        return '';
                    }}''', wanted_norm)
                    if ticked:
                        log_cb(f'  ✓ Đã tick: {ticked}', 'ok' if 'OK' in ticked else 'warn')
                        ticked_texts.append(wanted)
                    else:
                        log_cb(f'  ✗ Không tìm thấy checkbox: "{wanted}"', 'warn')
                except Exception as e:
                    log_cb(f'  ⚠ Lỗi tick "{wanted}": {e}', 'warn')
                page.wait_for_timeout(300)

            if len(ticked_texts) < len(CHECK_TEXTS):
                try:
                    ss = str(Path(output_dir) / f'debug_popup_{carrier or "all"}_batch{batch_num}.png')
                    page.screenshot(path=ss)
                    log_cb(f'  📸 Screenshot popup: {ss}', 'dim')
                except Exception:
                    pass
            page.wait_for_timeout(500)

            # ── 10. Create documents + network interception ──
            new_page_ref = []
            downloaded_files = []

            def _awb_on_new_page(p):
                log_cb(f'  🔔 Tab awbprint: {p.url[:120] if p.url else "(loading)"}', 'dim')
                new_page_ref.append(p)
                def _on_response(response):
                    if downloaded_files: return
                    ct = (response.headers.get('content-type', '') or '').lower()
                    url_lower = (response.url or '').lower()
                    if (('application/pdf' in ct or 'application/force-download' in ct or url_lower.endswith('.pdf') or 'download_sd_job' in url_lower) and response.ok):
                        try:
                            body = response.body()
                            if body and len(body) > 5000:
                                bp = str(Path(output_dir) / f'Shopee_{carrier or "all"}_batch{batch_num}_{datetime.now().strftime("%m-%d_%H-%M-%S")}.pdf')
                                with open(bp, 'wb') as f: f.write(body)
                                downloaded_files.append(bp)
                                log_cb(f'  💾 Network PDF: {Path(bp).name} ({len(body)//1024}KB)', 'ok')
                        except Exception as e:
                            log_cb(f'  ⚠ Response body error: {e}', 'warn')
                p.on('response', _on_response)
                def _awb_dl(dl):
                    if downloaded_files: return
                    bp = str(Path(output_dir) / (dl.suggested_filename or f'Shopee_{carrier or "all"}_batch{batch_num}_{datetime.now().strftime("%m-%d_%H-%M-%S")}.pdf'))
                    try:
                        dl.save_as(bp)
                        downloaded_files.append(bp)
                        log_cb(f'  💾 Download: {Path(bp).name}', 'ok')
                    except Exception as e:
                        log_cb(f'  ⚠ Save error: {e}', 'warn')
                p.on('download', _awb_dl)

            context.on('page', _awb_on_new_page)

            create_clicked = False
            for btn_text in ['Tạo phiếu đã chọn', 'Tạo phiếu', 'Tạo']:
                try:
                    btn = page.locator(f'button:has-text("{btn_text}")').first
                    if btn.count() > 0 and btn.is_visible(timeout=2000):
                        btn.click(force=True, timeout=5000)
                        log_cb(f'  ✓ Đã bấm "{btn_text}"', 'dim')
                        create_clicked = True
                        break
                except Exception:
                    pass
            if not create_clicked:
                try:
                    clicked = page.evaluate('''() => {
                        const btns = document.querySelectorAll('button');
                        for (const btn of btns) {
                            const txt = btn.textContent.trim();
                            if ((txt.includes('Tạo phiếu') || txt.includes('Tạo')) && btn.offsetParent !== null) { btn.click(); return txt; }
                        }
                        return '';
                    }''')
                    if clicked:
                        log_cb(f'  ✓ Fallback: đã bấm "{clicked}"', 'dim')
                        create_clicked = True
                except Exception:
                    pass
            if not create_clicked:
                log_cb('  ⚠ Không tìm thấy nút "Tạo phiếu đã chọn"', 'warn')
                try: context.remove_listener('page', _awb_on_new_page)
                except: pass
                continue

            log_cb('  ⏳ Đợi tab awbprint + PDF (tối đa 120 giây)...', 'dim')
            for _wait_i in range(60):
                if downloaded_files: break
                if stop_event and stop_event.is_set():
                    log_cb('  ⏹ Đã dừng theo yêu cầu', 'warn')
                    try: context.remove_listener('page', _awb_on_new_page)
                    except: pass
                    return all_batch_pdfs if all_batch_pdfs else pdf_files, playwright, browser, carrier_counts
                if not new_page_ref:
                    for p in context.pages:
                        if p != page and not p.is_closed():
                            _awb_on_new_page(p)
                            break
                if new_page_ref:
                    awb_page = new_page_ref[0]
                    if not awb_page.is_closed():
                        try:
                            if _wait_i == 5:
                                try: awb_page.wait_for_load_state('networkidle', timeout=30000)
                                except: pass
                            if _wait_i >= 10 and not downloaded_files:
                                for btn_text in ['In', 'In phiếu', 'Print', 'Tải về', 'Download', 'Lưu']:
                                    try:
                                        btn = awb_page.locator(f'button:has-text("{btn_text}")').first
                                        if btn.count() > 0 and btn.is_visible(timeout=500):
                                            btn.click(force=True, timeout=3000)
                                            log_cb(f'  ✓ Đã bấm "{btn_text}" trên tab awbprint', 'dim')
                                            break
                                    except: pass
                        except: pass
                page.wait_for_timeout(2000)

            try: context.remove_listener('page', _awb_on_new_page)
            except: pass

            if downloaded_files:
                log_cb(f'  ✅ Đã lấy {len(downloaded_files)} file PDF', 'ok')
            else:
                log_cb('  ⚠ Network không bắt được PDF — thử page.pdf()...', 'warn')
                try:
                    ss = str(Path(output_dir) / f'debug_no_dl_{carrier or "all"}_batch{batch_num}.png')
                    page.screenshot(path=ss)
                    log_cb(f'  📸 Screenshot: {ss}', 'dim')
                except: pass
                awb_page = new_page_ref[0] if new_page_ref else None
                if awb_page and not awb_page.is_closed():
                    try: awb_page.wait_for_load_state('networkidle', timeout=15000)
                    except: pass
                    awb_page.wait_for_timeout(5000)
                    try:
                        save_path = str(Path(output_dir) / f'Shopee_{carrier or "all"}_batch{batch_num}_{datetime.now().strftime("%m-%d_%H-%M-%S")}.pdf')
                        awb_page.pdf(path=save_path)
                        downloaded_files.append(save_path)
                        log_cb(f'  💾 page.pdf(): {Path(save_path).name}', 'ok')
                    except Exception as e2:
                        log_cb(f'  ⚠ Không lưu được PDF: {e2}', 'warn')

            for p in new_page_ref:
                try:
                    if not p.is_closed():
                        p.close()
                        log_cb('  ✓ Đã đóng tab awbprint', 'dim')
                except: pass

            page.wait_for_timeout(500)
            try: page.bring_to_front()
            except: pass
            page.wait_for_timeout(500)

            for f in downloaded_files:
                all_batch_pdfs.append(f)
                pdf_files.append(f)

            log_cb(f'  ✅ Batch {batch_num} hoàn thành ({len(downloaded_files)} file)', 'ok')
            page.wait_for_timeout(2000)

        if all_batch_pdfs:
            log_cb(f'📥 Tổng cộng {len(all_batch_pdfs)} file PDF sau {batch_num} batch', 'ok')
        return all_batch_pdfs if all_batch_pdfs else pdf_files, playwright, browser, carrier_counts

    except Exception as e:

        log_cb(f'  ✗ Lỗi: {e}', 'err')
        try:
            page.screenshot(path=str(Path(output_dir) / f'error_{carrier or "all"}.png'))
        except Exception:
            pass
        return pdf_files, playwright, browser, carrier_counts


def run_calculator(pdf_paths, output_dir, master_path, retail_path, template_path, log_cb, carrier=''):
    # Lazy import: tránh kéo pdfplumber + openpyxl + win32com (~6s) lúc khởi động app
    try:
        process_all, _, _, _ = _get_calculator()
    except ImportError:
        log_cb('✗ Calculator không khả dụng (thiếu module calculator)', 'err'); return []
    if not Path(master_path).exists(): log_cb(f'✗ Không tìm thấy master_data: {master_path}', 'err'); return []
    if not Path(template_path).exists(): log_cb(f'✗ Không tìm thấy template: {template_path}', 'err'); return []
    out_dir = str(output_dir)
    for p in pdf_paths:
        shutil.copy2(p, str(UPLOAD_DIR / Path(p).name))
    try:
        results = process_all(pdf_paths, out_dir, master_path, retail_path, carrier, template_path=template_path)
        for r in results:
            log_cb(f'  ✓ {r["rows"]} dòng | Qty={r["tong_qty"]} | Sold={r["tong_sold"]} | Promo={r["tong_promo"]}', 'ok')
            for key, fb in r['files'].items():
                src, dst = Path(fb), Path(out_dir) / Path(fb).name
                if src != dst and src.exists(): shutil.copy2(str(src), str(dst)); r['files'][key] = str(dst)
        return results
    except Exception as e: log_cb(f'  ✗ Lỗi: {e}', 'err'); return []

# ═══════════════════════════════════════════════════════════════
# WORKER & PRINTING FUNCTIONS
# ═══════════════════════════════════════════════════════════════
class AutomationWorker(QObject):
    log_message = Signal(str, str)
    state_changed = Signal(str, str)
    job_completed = Signal(object)
    result_file = Signal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._playwright = None
        self._browser = None
        self._stop_event = threading.Event()
        self._running = False

    @Slot(dict)
    def start_job(self, config: dict):
        # ── Dọn dẹp browser cũ từ job trước (nếu có) ──
        if self._browser is not None:
            try:
                self._browser.close()
            except Exception:
                pass
            self._browser = None
        if self._playwright is not None:
            import time as _time_cleanup
            _time_cleanup.sleep(0.3)
            try:
                self._playwright.stop()
            except Exception:
                pass
            self._playwright = None

        self._running = True
        self._stop_event.clear()

        # ── Khởi tạo ──
        all_pdf_paths = []
        all_results = []
        completed: set[str] = set()
        total_processed = 0

        pw = None
        br = None

        try:
            cookie = config['cookie']
            out_dir = config['output_dir']  # Đã có sẵn date subfolder từ main thread
            master = config['master']
            retail = config['retail']
            template = config['template']
            auto_print = config['auto_print']
            printer = config['printer']
            test_mode = config['test_mode']
            exclude_pre_orders = config.get('exclude_pre_orders', True)
            batch_size = config.get('batch_size', 0)
            pdf_settings = config.get('pdf_settings', 'paper=A4')
            carriers = config['carriers']

            os.makedirs(out_dir, exist_ok=True)

            # ── Quét 1 lần, skip carrier 0 đơn ──
            known_carrier_counts: dict[str, int] = {}

            for carrier, count in carriers:
                # ── Skip completed carriers khi resume ──
                if carrier in completed:
                    self.log_message.emit('info', f'⏭ Bỏ qua [{carrier}] — đã hoàn thành trước đó')
                    continue

                # ── Skip carriers đã scan thấy 0 đơn ──
                if carrier in known_carrier_counts and known_carrier_counts[carrier] == 0:
                    self.log_message.emit('dim', f'⏭ Bỏ qua [{carrier}] — đã quét thấy 0 đơn')
                    completed.add(carrier)
                    continue

                if self._stop_event.is_set():
                    self.log_message.emit('warn', 'Đã dừng theo yêu cầu.')
                    break

                carrier_display = carrier if carrier else 'tất cả'
                count_display = f'{count}' if count > 0 else 'tất cả'
                self.log_message.emit('info', f'📥 Tải PDF [{carrier_display}] ({count_display} đơn)...')

                pdf_paths, pw2, br2, scanned_counts = run_automation(
                    cookie, out_dir, count,
                    lambda m, t='': self.log_message.emit(t, m),
                    lambda s, m: self.state_changed.emit(s, m),
                    self._stop_event,
                    existing_playwright=pw, existing_browser=br,
                    carrier=carrier, test_mode=test_mode,
                    exclude_pre_orders=exclude_pre_orders)

                # ── Cập nhật known_carrier_counts từ lần scan đầu tiên ──
                if scanned_counts and not known_carrier_counts:
                    known_carrier_counts = scanned_counts
                    active = {k: v for k, v in scanned_counts.items() if v > 0}
                    self.log_message.emit('info', f'📊 Quét xong: {active if active else "(không có đơn nào)"}')

                if pw2: pw = pw2
                if br2: br = br2
                self._playwright = pw
                self._browser = br

                for p in pdf_paths:
                    if Path(p).exists():
                        self.log_message.emit('info', f'📄 {Path(p).name}')
                        self.result_file.emit(str(p))
                all_pdf_paths.extend(pdf_paths)
                tag = 'ok' if pdf_paths else 'warn'
                self.log_message.emit(tag, f'📥 [{carrier_display}]: đã tải {len(pdf_paths)} file PDF')

                carrier_results = []
                if pdf_paths:
                    # ── Bước 1: Tách shipping label từ PDF gộp ──
                    _, _, _, detect_fn, split_fn = _get_calculator()
                    shipping_files = []
                    for p in pdf_paths:
                        pdf_type = detect_fn(p)
                        if pdf_type == 'shopee':
                            picking_path, shipping_path, _ = split_fn(p, out_dir)
                            if shipping_path:
                                shipping_files.append((p, shipping_path))

                    # ── Bước 2: In shipping label TRƯỚC ──
                    if auto_print and shipping_files:
                        self.log_message.emit('info', f'🖨️ [{carrier_display}]: In shipping label...')
                        for orig_p, shipping_f in shipping_files:
                            if Path(shipping_f).exists():
                                try:
                                    _print_file(str(shipping_f), printer, pdf_settings=pdf_settings, batch_size=batch_size,
                                                log_cb=lambda m, t='': self.log_message.emit(t, m))
                                    self.log_message.emit('ok', f'  ✓ Đã gửi in: {Path(shipping_f).name}')
                                except Exception as e:
                                    self.log_message.emit('err', f'  ✗ Lỗi in {Path(shipping_f).name}: {e}')

                    # ── Bước 3: Tính bill ──
                    self.log_message.emit('info', f'📊 Đang tính bill [{carrier_display}]...')
                    carrier_results = run_calculator(
                        pdf_paths, out_dir, master, retail, template,
                        lambda m, t='': self.log_message.emit(t, m),
                        carrier=carrier)
                    for r in carrier_results:
                        for key, lbl in [('xlsx_report', '📊')]:
                            fp = r['files'].get(key)
                            if fp and Path(fp).exists():
                                self.log_message.emit('info', f'{lbl} {Path(fp).name}')
                                self.result_file.emit(f'{lbl} {Path(fp).name}')
                    all_results.extend(carrier_results)

                # ── Dọn file shipping label (đã in hoặc không cần in) ──
                for p in pdf_paths:
                    stem = Path(p).stem
                    shipping_f = Path(out_dir) / f'{stem}_shipping_label.pdf'
                    if shipping_f.exists():
                        try: shipping_f.unlink()
                        except: pass

                # In báo cáo sau khi tính toán
                if auto_print and carrier_results:
                    for r in carrier_results:
                        fp = r['files'].get('pdf_report') or r['files'].get('xlsx_report')
                        if fp and Path(fp).exists():
                            try:
                                for copy_num in [1, 2]:
                                    self.log_message.emit('info', f'  🖨️ In bản {copy_num}/2: {Path(fp).name}')
                                    _print_file(fp, printer, pdf_settings=pdf_settings, batch_size=batch_size,
                                                log_cb=lambda m, t='': self.log_message.emit(t, m))
                                    if copy_num == 1:
                                        import time as _t3; _t3.sleep(2)
                                self.log_message.emit('ok', f'  ✓ Đã in báo cáo 2 bản: {Path(fp).name}')
                            except Exception as e:
                                self.log_message.emit('err', f'  ✗ Lỗi in báo cáo: {e}')

                # ── Đánh dấu carrier hoàn thành ──
                completed.add(carrier)
                total_processed += count if count > 0 else 0

            # ── Tất cả carriers hoàn thành ──
            self.log_message.emit('bold_ok', '🏁 HOÀN THÀNH!')
            self.state_changed.emit('done', f'✅ Hoàn thành lúc {datetime.now().strftime("%H:%M:%S")}')
            try:
                if br: br.close()
            except Exception: pass
            try:
                if pw: pw.stop()
            except Exception: pass
            self._playwright = None
            self._browser = None
            self.job_completed.emit({
                'playwright': None, 'browser': None,
                'pdf_paths': all_pdf_paths, 'results': all_results,
                'output_dir': out_dir,
            })
        except Exception as e:
            self.log_message.emit('err', f'✗ Lỗi: {e}')
            self.state_changed.emit('error', f'✗ {e}')
            try:
                if self._browser: self._browser.close()
            except Exception: pass
            try:
                if self._playwright: self._playwright.stop()
            except Exception: pass
            self._playwright = None
            self._browser = None
            self.job_completed.emit({})
        finally:
            self._running = False

    @Slot()
    def stop_job(self): self._stop_event.set()

    @Slot()
    def shutdown(self):
        self._stop_event.set()
        try:
            if self._browser: self._browser.close()
        except: pass
        try:
            if self._playwright: self._playwright.stop()
        except: pass

_foxit_exe_cache = None


def _find_foxit_exe():
    """Tìm Foxit PDF Reader — dùng XPS Print Path, spool nhẹ ~15MB."""
    global _foxit_exe_cache
    if _foxit_exe_cache is not None:
        return _foxit_exe_cache or ''
    foxit_paths = [
        r'C:\Program Files (x86)\Foxit Software\Foxit PDF Reader\FoxitPDFReader.exe',
        r'C:\Program Files\Foxit Software\Foxit PDF Reader\FoxitPDFReader.exe',
    ]
    for fp in foxit_paths:
        if Path(fp).exists():
            _foxit_exe_cache = fp
            return fp
    _foxit_exe_cache = ''
    return ''


def _check_print_errors(printer_name, doc_name_hint='', log_cb=None, timeout=60):
    """Poll print queue để phát hiện lỗi máy in (kẹt giấy, hết mực...).
    Đợi đến khi job Complete hoặc Error thì trả về.
    timeout: số giây tối đa chờ (mặc định 60s).
    Lưu ý: nếu job chưa từng xuất hiện trong queue sau 15s → coi như đã in xong quá nhanh."""
    import subprocess as _sp, os as _os
    deadline = __import__('time').time() + timeout
    last_status = ''
    error_reported = False
    never_seen_deadline = __import__('time').time() + 15  # 15s đầu phải thấy job, nếu không → exit sớm
    while __import__('time').time() < deadline:
        __import__('time').sleep(5)
        result = _sp.run(['powershell', '-NoProfile', '-WindowStyle', 'Hidden', '-Command',
            f"$jobs = Get-PrintJob -PrinterName '{printer_name}' -ErrorAction SilentlyContinue"
            + (f" | Where-Object {{ $_.DocumentName -like '*{doc_name_hint[:30]}*' }}" if doc_name_hint else "")
            + " | Select-Object JobStatus | ConvertTo-Json -Compress"
        ], capture_output=True, text=True,
           creationflags=_sp.CREATE_NO_WINDOW if sys.platform == 'win32' else 0)
        status_raw = result.stdout.strip()
        status = ''
        if status_raw:
            try:
                import json as _json
                job_info = _json.loads(status_raw)
                if isinstance(job_info, list):
                    job_info = job_info[0] if job_info else {}
                status = job_info.get('JobStatus', '')
            except Exception:
                status = status_raw

        if status and status != last_status:
            if log_cb: log_cb(f'  🖨️ Trạng thái: {status}', 'dim')
            last_status = status

        if status:
            if 'Complete' in status and 'Printing' not in status and 'Spooling' not in status:
                if log_cb: log_cb(f'  ✅ Job in đã hoàn thành', 'dim')
                return True

            error_msgs = {
                'PaperJam': '🛑 KẸT GIẤY! Hãy gỡ giấy kẹt rồi nhấn nút trên máy in.',
                'PaperOut': '📄 HẾT GIẤY! Hãy nạp thêm giấy vào khay.',
                'TonerLow': '⚠ SẮP HẾT MỰC! Chuẩn bị thay mực.',
                'NoToner': '🖌 HẾT MỰC! Cần thay cartridge mực.',
                'Offline': '🔌 MÁY IN MẤT KẾT NỐI! Kiểm tra cáp/WiFi.',
                'Paused': '⏸ MÁY IN ĐANG TẠM DỪNG! Kiểm tra nút trên máy.',
                'DoorOpen': '🚪 NẮP MÁY IN ĐANG MỞ! Đóng nắp lại.',
                'OutputFull': '📦 KHAY RA ĐẦY! Lấy giấy đã in ra.',
            }
            for code, msg in error_msgs.items():
                if code in status and not error_reported:
                    if log_cb: log_cb(f'  {msg}', 'err')
                    error_reported = True
                    break

            if 'Error' in status and not error_reported:
                if log_cb: log_cb(f'  ⚠ Lỗi máy in: {status}', 'err')
                return False
        else:
            # Job đã biến mất khỏi queue (đã in xong và được xóa)
            if last_status:
                if log_cb: log_cb(f'  ✅ Job in đã rời queue (đã in xong)', 'dim')
                return True
            # Neu da qua 15s ma chua bao gio thay job → may in xu ly qua nhanh, coi nhu xong
            if __import__('time').time() > never_seen_deadline:
                if log_cb: log_cb(f'  ⚡ Job in qua nhanh, khong thay trong queue — coi nhu da in xong', 'dim')
                return True

    # Hết timeout — có thể job vẫn đang in, không chặn tiến trình
    if log_cb: log_cb(f'  ⚠ Hết {timeout}s chờ — tiếp tục (job có thể vẫn đang in)', 'warn')
    return True


def _wait_print_queue(printer_name, max_jobs=2, timeout=30):
    """Đợi hàng đợi máy in ≤ max_jobs rồi mới gửi job mới (tránh quá tải RAM máy in).
    timeout: số giây tối đa chờ (mặc định 30s)."""
    import subprocess as _sp, time as _t
    deadline = _t.time() + timeout
    waited = False
    while _t.time() < deadline:
        result = _sp.run(['powershell', '-NoProfile', '-WindowStyle', 'Hidden', '-Command',
            f"@(Get-PrintJob -PrinterName '{printer_name}' -ErrorAction SilentlyContinue).Count"
        ], capture_output=True, text=True, creationflags=_sp.CREATE_NO_WINDOW if sys.platform == 'win32' else 0)
        try:
            count = int(result.stdout.strip())
        except ValueError:
            count = 0
        if count <= max_jobs:
            if waited:
                print(f'   ✓ Hàng đợi đã trống (sau ~{int(_t.time() - (deadline - timeout))}s)')
            return
        if not waited:
            waited = True
            print(f'   ⏳ Hàng đợi có {count} job(s) — đợi giảm xuống ≤{max_jobs}...')
        _t.sleep(1)  # kiểm tra mỗi 1 giây
    # Nếu hết timeout mà vẫn còn job → log cảnh báo và in tiếp (tránh treo vĩnh viễn)
    print(f'   ⚠ Hết {timeout}s chờ — hàng đợi vẫn còn job, in tiếp...')


def _print_file(file_path, printer_name, pdf_settings='paper=A4', log_cb=None, batch_size=0):
    import subprocess, os as _os
    fp = str(file_path)

    try:
        if fp.lower().endswith('.pdf'):
            foxit_exe = _find_foxit_exe()
            if not foxit_exe:
                raise RuntimeError(
                    'Không tìm thấy Foxit PDF Reader. '
                    'Vui lòng cài Foxit PDF Reader để in file PDF.'
                )

            print_path = fp
            temp_merged = None
            # Chỉ merge 2-up file shipping label, không merge file báo cáo
            fname_lower = Path(fp).name.lower()
            if 'shipping' in fname_lower or 'vận chuyển' in fname_lower:
                try:
                    if log_cb: log_cb(f'  📐 Đang merge 2-up: {Path(fp).name}...', 'dim')
                    temp_merged = _merge_pdf_2up(fp)
                    if temp_merged: print_path = temp_merged
                    if log_cb: log_cb(f'  ✓ Merge 2-up hoàn tất', 'dim')
                except Exception as e:
                    if log_cb: log_cb(f'  ⚠ Merge 2-up lỗi ({e}) — in file gốc', 'warn')

            if foxit_exe:
                # ── Đọc số trang để quyết định batch splitting ──
                try:
                    from pypdf import PdfReader as _PdfReader, PdfWriter as _PdfWriter
                    reader = _PdfReader(print_path)
                    total_pages = len(reader.pages)
                except Exception:
                    reader = None
                    total_pages = 0

                # ── Batch splitting ──
                if batch_size > 0 and reader and total_pages > batch_size:
                    total_batches = (total_pages + batch_size - 1) // batch_size
                    if log_cb: log_cb(f'  📦 Foxit: Chia {total_pages} tờ → {total_batches} batch ({batch_size} tờ/batch)', 'info')
                    batch_num = 0
                    for start in range(0, total_pages, batch_size):
                        batch_num += 1
                        end = min(start + batch_size, total_pages)
                        if log_cb: log_cb(f'  🖨️ Batch {batch_num}/{total_batches} (tờ {start+1}-{end})...', 'info')
                        batch_writer = _PdfWriter()
                        for i in range(start, end):
                            batch_writer.add_page(reader.pages[i])
                        batch_path = print_path + f'.batch{batch_num}.pdf'
                        with open(batch_path, 'wb') as bf:
                            batch_writer.write(bf)

                        if log_cb: log_cb(f'  ⏳ Đợi hàng đợi máy in trống...', 'dim')
                        _wait_print_queue(printer_name, max_jobs=0)
                        if log_cb: log_cb(f'  ▶ Gửi lệnh in qua Foxit (batch {batch_num})...', 'info')
                        cmd = [foxit_exe, '/t', batch_path, printer_name]
                        result = subprocess.run(cmd, check=False, timeout=3600)
                        if log_cb: log_cb(f'  ✓ Foxit batch {batch_num} đã thoát (exit code: {result.returncode})', 'dim')
                        if result.returncode != 0:
                            raise RuntimeError(f'Foxit batch {batch_num} exit code: {result.returncode}')
                        if log_cb: log_cb(f'  🔍 Đang kiểm tra trạng thái in batch {batch_num}...', 'dim')
                        _check_print_errors(printer_name, doc_name_hint=os.path.basename(batch_path), log_cb=log_cb)

                        if log_cb: log_cb(f'  ✅ Batch {batch_num}/{total_batches} đã in xong', 'ok')
                        try: _os.remove(batch_path)
                        except: pass

                else:
                    # ── In thẳng không batch ──
                    if log_cb: log_cb(f'  ⏳ Đợi hàng đợi máy in trống (max_jobs=0)...', 'dim')
                    _wait_print_queue(printer_name, max_jobs=0)
                    if log_cb: log_cb(f'  ▶ Gửi lệnh in qua Foxit: {os.path.basename(print_path)}', 'info')
                    cmd = [foxit_exe, '/t', print_path, printer_name]
                    result = subprocess.run(cmd, check=False, timeout=3600)
                    if log_cb: log_cb(f'  ✓ Foxit đã thoát (exit code: {result.returncode})', 'dim')
                    if result.returncode != 0:
                        raise RuntimeError(f'Foxit exit code: {result.returncode}')
                    if log_cb: log_cb(f'  🔍 Đang kiểm tra trạng thái in...', 'dim')
                    _check_print_errors(printer_name, doc_name_hint=os.path.basename(print_path), log_cb=log_cb)
                    if log_cb: log_cb(f'  ✅ In hoàn tất', 'ok')
            if temp_merged:
                def _cleanup(p=temp_merged):
                    import time; time.sleep(5)
                    try: _os.remove(p)
                    except: pass
                threading.Timer(5, _cleanup).start()
            return

        if fp.lower().endswith('.xlsx') or fp.lower().endswith('.xls'):
            try:
                import pythoncom, win32com.client, time as _t_excel
                # Đợi queue trống + delay cứng để đảm bảo 2 job không bị gộp
                _wait_print_queue(printer_name)
                _t_excel.sleep(1)
                pythoncom.CoInitialize()
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                workbook = excel.Workbooks.Open(_os.path.abspath(fp))
                workbook.PrintOut(ActivePrinter=printer_name, FitToPagesWide=1, FitToPagesTall=False)
                # Đợi Excel spool xong job ra queue rồi mới đóng
                _t_excel.sleep(2)
                workbook.Close(False); excel.Quit(); pythoncom.CoUninitialize()
                # Đợi job đã chắc chắn vào queue
                _t_excel.sleep(1)
                return
            except Exception:
                try: pythoncom.CoUninitialize()
                except: pass
            raise RuntimeError(
                'Không thể in file Excel. Máy cần cài Microsoft Excel.'
            )
    except Exception:
        raise

def _merge_pdf_2up(pdf_path):
    import os as _os
    try: from pypdf import PdfReader, PdfWriter, PageObject, Transformation
    except ImportError: return None
    try:
        reader = PdfReader(pdf_path)
        if len(reader.pages) < 1: return None
        canvas_w, canvas_h = 842, 595  # A4 ngang
        margin_left, margin_top, gap = 5, 5, 0
        writer = PdfWriter()
        avail_w = canvas_w - 2 * margin_left - gap
        half_w = avail_w / 2
        avail_h = canvas_h - 2 * margin_top
        for pair_start in range(0, len(reader.pages), 2):
            pair = reader.pages[pair_start:pair_start+2]
            canvas = PageObject.create_blank_page(width=canvas_w, height=canvas_h)
            for i, page in enumerate(pair):
                pw = float(page.mediabox.width)
                ph = float(page.mediabox.height)
                # Tính scale vừa khít
                scale = 0.7
                sw, sh = pw * scale, ph * scale
                tx = margin_left + i * (half_w + gap)
                ty = canvas_h - margin_top - sh
                canvas.merge_transformed_page(page, Transformation().scale(scale).translate(tx / scale, ty / scale))
            writer.add_page(canvas)
        # Nén để giảm dung lượng
        for page in writer.pages:
            page.compress_content_streams()
        temp_path = pdf_path + '.2up.pdf'
        with open(temp_path, 'wb') as f:
            writer.write(f)
        return temp_path
    except:
        return None

def _get_printers() -> list:
    printers = []
    try:
        import win32print
        for info in win32print.EnumPrinters(win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS, None, 1):
            name = info[2].strip() if info[2] else ''
            if name: printers.append(name)
    except: pass
    if not printers:
        try:
            import subprocess
            result = subprocess.run(['powershell', '-Command', "Get-Printer | Select-Object -ExpandProperty Name | Where-Object { $_ -notlike '*Microsoft*' -and $_ -notlike '*Fax*' -and $_ -notlike '*OneNote*' -and $_ -notlike '*XPS*' }"], capture_output=True, text=True, timeout=10)
            for line in result.stdout.strip().split('\n'):
                name = line.strip()
                if name and name not in printers: printers.append(name)
        except: pass
    return printers


def _get_default_printer() -> str:
    """Lấy tên máy in mặc định của hệ thống."""
    try:
        import win32print
        return win32print.GetDefaultPrinter()
    except Exception:
        pass
    try:
        import subprocess
        result = subprocess.run(
            ['powershell', '-NoProfile', '-Command',
             '(Get-WmiObject -Query "SELECT * FROM Win32_Printer WHERE Default=True").Name'],
            capture_output=True, text=True, timeout=5,
            creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == 'win32' else 0)
        name = result.stdout.strip()
        if name:
            return name
    except Exception:
        pass
    return ''


# ═══════════════════════════════════════════════════════════════
# WIDGETS
# ═══════════════════════════════════════════════════════════════
class FileRowWidget(QWidget):
    path_changed = Signal(str)

    def __init__(self, label_text: str, file_filter: str = '', is_dir: bool = False, parent=None):
        super().__init__(parent)
        self._real_path = ''
        self._filter = file_filter
        self._is_dir = is_dir

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(2)

        row = QHBoxLayout()
        row.setSpacing(6)

        lbl = QLabel(label_text)
        lbl.setFixedWidth(130)
        lbl.setStyleSheet('font-weight: 600; color: #1E293B; font-size: 10pt;')
        row.addWidget(lbl)

        self.path_edit = QLineEdit()
        self.path_edit.setReadOnly(True)
        self.path_edit.setObjectName('pathEdit')
        self.path_edit.setPlaceholderText('Chưa chọn...')
        row.addWidget(self.path_edit, 1)

        btn = QPushButton('Chọn')
        btn.setObjectName('browseBtn')
        btn.setCursor(Qt.PointingHandCursor)
        btn.clicked.connect(self._browse)
        row.addWidget(btn)

        layout.addLayout(row)

        self.status_label = QLabel('❌ Chưa chọn')
        self.status_label.setObjectName('statusLabel')
        self.status_label.setStyleSheet('font-size: 8.5pt; color: #DC2626; padding-left: 136px;')
        layout.addWidget(self.status_label)

    def _browse(self):
        if self._is_dir:
            p = QFileDialog.getExistingDirectory(self, 'Chọn thư mục')
            if p:
                self._real_path = p
                self.path_edit.setText(p)
                self._update_status_dir()
                self.path_changed.emit(p)
        else:
            p, _ = QFileDialog.getOpenFileName(self, 'Chọn file', '', self._filter)
            if p:
                self._real_path = p
                self.path_edit.setText(p)
                self.path_changed.emit(p)

    def set_path(self, path: str):
        self._real_path = path
        self.path_edit.setText(path)

    def get_real_path(self) -> str:
        return self._real_path

    def _update_status_dir(self):
        if self._real_path and Path(self._real_path).exists():
            self.status_label.setText('✅ Đã cập nhật thư mục lưu')
            self.status_label.setStyleSheet('font-size: 8.5pt; color: #059669; padding-left: 136px;')
        else:
            self.status_label.setText('❌ Chưa chọn')
            self.status_label.setStyleSheet('font-size: 8.5pt; color: #DC2626; padding-left: 136px;')

    def update_cookie_status(self):
        p = Path(self._real_path) if self._real_path else Path('')
        if self._real_path and p.exists():
            try:
                d = json.loads(p.read_text(encoding='utf-8'))
                cs = d.get('cookies', d)
                n = len(cs) if isinstance(cs, list) else 0
                self.status_label.setText(f'✅ Đã nạp {n} cookies')
                self.status_label.setStyleSheet('font-size: 8.5pt; color: #059669; padding-left: 136px;')
            except Exception:
                self.status_label.setText('⚠ File bị lỗi hoặc sai định dạng')
                self.status_label.setStyleSheet('font-size: 8.5pt; color: #D97706; padding-left: 136px;')
        else:
            self.status_label.setText('❌ Chưa chọn file hợp lệ')
            self.status_label.setStyleSheet('font-size: 8.5pt; color: #DC2626; padding-left: 136px;')

    def update_excel_status(self):
        p = Path(self._real_path) if self._real_path else Path('')
        if self._real_path and p.exists() and p.suffix.lower() == '.xlsx':
            try:
                from openpyxl import load_workbook
                wb = load_workbook(str(p), data_only=True, read_only=True)
                sheet = wb.active
                n = sheet.max_row - 1 if sheet.max_row else 0
                wb.close()
                self.status_label.setText(f'✅ Đã nạp {n} SKU')
                self.status_label.setStyleSheet('font-size: 8.5pt; color: #059669; padding-left: 136px;')
            except Exception:
                self.status_label.setText('⚠ File bị lỗi')
                self.status_label.setStyleSheet('font-size: 8.5pt; color: #D97706; padding-left: 136px;')
        else:
            self.status_label.setText('❌ Chưa chọn file hợp lệ')
            self.status_label.setStyleSheet('font-size: 8.5pt; color: #DC2626; padding-left: 136px;')


# ============================================================
# MAIN APP — QMainWindow with PySide6 GUI (SaaS Light Theme)
# ============================================================
class App(QMainWindow):
    trigger_job = Signal(dict)
    _test_log = Signal(str, str)  # signal cho test tab (thread-safe)

    TAG_COLORS = {
        "ts": "#64748B", "ok": "#10B981", "err": "#F87171",
        "warn": "#FBBF24", "info": "#60A5FA", "batch": "#C084FC",
        "result": "#F472B6", "dim": "#64748B", "bold_ok": "#10B981",
    }

    def __init__(self):
        super().__init__()
        self.setWindowTitle("ShopeePrint")
        self.resize(850, 850)
        self.setMinimumSize(600, 700)

        self._cookie_real = str(DEFAULT_COOKIE) if DEFAULT_COOKIE.exists() else ""
        local_master = BASE_DIR / "mã combo.xlsx"
        if local_master.exists(): self._master_real = str(local_master)
        elif MASTER_DEFAULT.exists(): self._master_real = str(MASTER_DEFAULT)
        else: self._master_real = ""
        local_retail = BASE_DIR / "sp bán lẻ.xlsx"
        if local_retail.exists(): self._retail_real = str(local_retail)
        elif RETAIL_DEFAULT.exists(): self._retail_real = str(RETAIL_DEFAULT)
        else: self._retail_real = ""
        local_template = BASE_DIR / "Bảng thống kê hàng.xlsx"
        if local_template.exists(): self._template_real = str(local_template)
        elif TEMPLATE_DEFAULT.exists(): self._template_real = str(TEMPLATE_DEFAULT)
        else: self._template_real = ""

        self.running = False
        self.scheduler_active = False
        self.result_files = []
        self._stop_event = threading.Event()

        self._sched_mode = "weekly"
        self._sched_interval_hours = 1
        self._sched_weekly_config: dict[int, list[tuple[int, int]]] = {}  # 0=Thứ 2..6=Chủ Nhật
        self._sched_next_run = None
        self._sched_last_run = None

        # Weekly scheduler widget refs (assigned in _build_schedule_tab)
        self.weekly_rb = None
        self.weekly_panel = None
        self.weekly_day_checkboxes: dict[int, QCheckBox] = {}
        self.weekly_day_time_edits: dict[int, QLineEdit] = {}
        self.weekly_master_time_edit = None

        self._worker_thread = QThread()
        self._worker = AutomationWorker()
        self._worker.moveToThread(self._worker_thread)
        self._worker.log_message.connect(self._on_log_message)
        self._worker.state_changed.connect(self._on_state_changed)
        self._worker.job_completed.connect(self._on_job_completed)
        self._worker.result_file.connect(self._add_result)
        self.trigger_job.connect(self._worker.start_job)
        self._test_log.connect(self._log_html)  # test tab log (thread-safe)
        self._worker_thread.start()

        self._sched_timer = QTimer(self)
        self._sched_timer.setInterval(1000)
        self._sched_timer.timeout.connect(self._check_schedule)

        self._apply_stylesheet()
        self._build_ui()

        self._update_cookie_status()
        self._update_master_status()
        self._update_retail_status()
        self._update_template_status()
        out_dir = self.output_row.get_real_path()
        if out_dir:
            os.makedirs(out_dir, exist_ok=True)
        self._load_preview_data()

    # ═══════════════════════════════════════════════════════
    # QSS — FIXED BLACK BACKGROUND BUG
    # ═══════════════════════════════════════════════════════
    def _apply_stylesheet(self):
        qss = """
            QMainWindow { background: #F8FAFC; }
            
            /* Fix Black Background Bug on ScrollArea and StackedWidget */
            QWidget#scrollContent { background: #F8FAFC; }
            QScrollArea, QStackedWidget { background: #F8FAFC; border: none; }
            
            /* Header */
            QWidget#headerBar { background: #EE4D2D; }
            QWidget#headerBar QLabel#headerTitle { color: #FFFFFF; font-size: 20px; font-weight: bold; }
            QWidget#headerBar QLabel#headerSubtitle { color: #FFD4C4; font-size: 12px; margin-top: 2px;}

            /* Tabs Navigation */
            QWidget#tabBar { background: #FFFFFF; border-bottom: 1px solid #E2E8F0; }
            QPushButton#tabBtn { background: transparent; color: #64748B; font-weight: bold; font-size: 14px; padding: 14px 24px; border: none; border-bottom: 3px solid transparent; }
            QPushButton#tabBtn:hover { color: #1E293B; border-bottom: 3px solid #CBD5E1; }
            QPushButton#tabBtn[active="true"] { color: #059669; border-bottom: 3px solid #059669; }

            /* GroupBox */
            QGroupBox { 
                background: #FFFFFF; 
                border: 1px solid #E2E8F0; 
                border-radius: 12px; 
                margin-top: 24px; 
                padding-top: 36px; 
                padding-left: 20px; 
                padding-right: 20px; 
                padding-bottom: 20px; 
                font-weight: bold; 
                font-size: 15px; 
                color: #1E293B; 
            }
            QGroupBox::title { subcontrol-origin: margin; subcontrol-position: top left; left: 20px; top: 0px; color: #1E293B; font-size: 15px; font-weight: bold; }

            /* Inputs */
            QLineEdit, QSpinBox, QComboBox { background: #F1F5F9; border: 1px solid #E2E8F0; border-radius: 6px; padding: 8px 12px; color: #1E293B; font-size: 14px; }
            QLineEdit:focus, QSpinBox:focus, QComboBox:focus { border-color: #059669; background: #FFFFFF; }
            
            /* Search Box specifically */
            QLineEdit#searchBox { border-radius: 16px; padding: 6px 14px; font-size: 13px; }

            QComboBox::drop-down { border: none; width: 24px; }
            QComboBox QAbstractItemView { background: #FFFFFF; border: 1px solid #E2E8F0; selection-background-color: #D1FAE5; selection-color: #1E293B; }

            /* Regular Buttons */
            QPushButton#browseBtn, QPushButton#refreshBtn { background: #FFFFFF; border: 1px solid #CBD5E1; border-radius: 6px; padding: 8px 16px; color: #1E293B; font-weight: bold; font-size: 14px; }
            QPushButton#browseBtn:hover, QPushButton#refreshBtn:hover { background: #F8FAFC; border-color: #94A3B8; }

            /* Action Buttons */
            QPushButton#schedBtn { background: #059669; color: #FFFFFF; border: none; border-radius: 8px; padding: 12px; font-weight: bold; font-size: 16px; }
            QPushButton#schedBtn:hover { background: #047857; }
            QPushButton#schedBtn:disabled { background: #A7F3D0; }

            QPushButton#stopBtn { background: #FEF2F2; color: #DC2626; border: 1px solid #FECACA; border-radius: 8px; padding: 12px; font-weight: bold; font-size: 14px; }
            QPushButton#stopBtn:hover { background: #FEE2E2; }
            QPushButton#stopBtn:disabled { background: #F8FAFC; color: #9CA3AF; border-color: #E2E8F0; }

            /* Small Icon Buttons */
            QPushButton#smallBtn { background: transparent; border: none; padding: 6px; color: #64748B; font-size: 12px; font-weight: bold; }
            QPushButton#smallBtn:hover { background: #F1F5F9; color: #1E293B; border-radius: 4px; }

            /* Preview Tabs */
            QPushButton#previewTabBtn { background: transparent; border: 1px solid transparent; color: #64748B; padding: 6px 12px; border-radius: 4px; font-size: 13px; font-weight: bold; }
            QPushButton#previewTabBtn[active="true"] { background: #F1F5F9; border: 1px solid #E2E8F0; color: #1E293B; }

            /* Tables */
            QTableWidget { background: #FFFFFF; border: 1px solid #E2E8F0; border-radius: 8px; gridline-color: #E2E8F0; font-size: 13px; color: #1E293B; }
            QTableWidget::item { padding: 4px; }
            QTableWidget::item:selected { background: #F8FAFC; color: #1E293B; }
            QTableWidget QLineEdit { padding: 1px 2px; background: #FFFFFF; border: 2px solid #059669; border-radius: 2px; color: #1E293B; font-size: 13px; }
            QHeaderView::section { background: #F1F5F9; color: #1E293B; font-weight: bold; padding: 10px 12px; border: none; border-right: 1px solid #E2E8F0; border-bottom: 1px solid #E2E8F0; font-size: 13px; }

            /* Checkboxes & Radio */
            QCheckBox, QRadioButton { color: #1E293B; font-size: 10pt; spacing: 8px; }
            QCheckBox::indicator, QRadioButton::indicator { width: 16px; height: 16px; }

            /* Log */
            QTextEdit#logView { background: #0F172A; color: #F8FAFC; border: 1px solid #E2E8F0; border-radius: 8px; font-family: "Consolas", monospace; font-size: 13px; padding: 12px; selection-background-color: #1E3A5F; selection-color: #FFFFFF; }

            /* Result List */
            QListWidget#resultList { background: #FFFFFF; border: 1px solid #E2E8F0; border-radius: 8px; padding: 4px; color: #1E293B; font-size: 13px; }
            QListWidget#resultList::item { padding: 8px; border-bottom: 1px solid #F1F5F9; }
            QListWidget#resultList::item:selected { background: #D1FAE5; color: #047857; font-weight: bold; border-radius: 4px; }

            /* Scrollbars */
            QScrollBar:vertical { background: transparent; width: 8px; }
            QScrollBar::handle:vertical { background: #CBD5E1; border-radius: 4px; min-height: 30px; }
            QScrollBar::handle:vertical:hover { background: #94A3B8; }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }

            /* Bottom Bar */
            QWidget#bottomBar { background: #FFFFFF; border-top: 1px solid #E2E8F0; }
            QLabel#statusLabel { font-weight: bold; font-size: 14px; padding: 4px 0; }
            QFrame#divider { background: #E2E8F0; max-height: 1px; margin: 12px 0; }
        """
        self.setStyleSheet(qss)
        font = QFont("Segoe UI", 10)
        self.setFont(font)

    # ═══════════════════════════════════════════════════════
    # BUILD UI
    # ═══════════════════════════════════════════════════════
    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        main_layout.addWidget(self._build_header())
        main_layout.addWidget(self._build_tab_bar())

        self.content_stack = QStackedWidget()
        self.content_stack.addWidget(self._build_files_tab())
        self.content_stack.addWidget(self._build_print_tab())
        self.content_stack.addWidget(self._build_schedule_tab())
        self.content_stack.addWidget(self._build_log_tab())
        self.content_stack.addWidget(self._build_test_tab())
        self.content_stack.addWidget(self._build_aggregate_tab())
        main_layout.addWidget(self.content_stack, 1)

        main_layout.addWidget(self._build_bottom_bar())

    def _build_header(self):
        h = QWidget()
        h.setObjectName("headerBar")
        h.setFixedHeight(70)
        layout = QVBoxLayout(h)
        layout.setContentsMargins(24, 12, 24, 12)
        layout.setSpacing(2)

        title = QLabel("ShopeePrint")
        title.setObjectName("headerTitle")
        layout.addWidget(title)

        subtitle = QLabel("Automation & Bill Calculate")
        subtitle.setObjectName("headerSubtitle")
        layout.addWidget(subtitle)

        return h

    def _build_tab_bar(self):
        bar = QWidget()
        bar.setObjectName("tabBar")
        layout = QHBoxLayout(bar)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        self.tab_buttons = []
        self.tab_button_group = QButtonGroup(self)
        self.tab_button_group.setExclusive(True)

        tab_labels = [
            "📦 Tệp dữ liệu",
            "🖨️ Cấu hình In",
            "⏰ Lịch trình",
            "📋 Nhật ký & Kết quả",
            "🧪 Test",
            "📊 Tổng hợp",
        ]

        for i, label in enumerate(tab_labels):
            btn = QPushButton(label)
            btn.setObjectName("tabBtn")
            btn.setCheckable(True)
            btn.setCursor(Qt.PointingHandCursor)
            btn.setProperty("tabIndex", i)
            if i == 0:
                btn.setChecked(True)
                btn.setProperty("active", True)
                btn.style().unpolish(btn)
                btn.style().polish(btn)
            self.tab_button_group.addButton(btn, i)
            self.tab_buttons.append(btn)
            layout.addWidget(btn)

        layout.addStretch()
        self.tab_button_group.idClicked.connect(self._on_tab_changed)
        return bar

    def _on_tab_changed(self, idx: int):
        self.content_stack.setCurrentIndex(idx)
        for i, btn in enumerate(self.tab_buttons):
            is_active = (i == idx)
            btn.setProperty("active", is_active)
            btn.style().unpolish(btn)
            btn.style().polish(btn)

    # ═══════════════════════════════════════════════════════
    # TAB 0: TỆP DỮ LIỆU & BẢNG PREVIEW
    # ═══════════════════════════════════════════════════════
    def _build_files_tab(self):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        
        w = QWidget()
        w.setObjectName("scrollContent")
        layout = QVBoxLayout(w)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(16)

        # ── Group 1: Khai báo đường dẫn ──
        gb1 = QGroupBox("Khai báo đường dẫn")
        gb1_layout = QVBoxLayout(gb1)
        gb1_layout.setSpacing(12)
        gb1_layout.setContentsMargins(0, 0, 0, 0) # Use QSS paddings

        self.cookie_row = FileRowWidget("🍪 Cookie (JSON)", "JSON Files (*.json)")
        self.cookie_row.set_path(self._cookie_real)
        self.cookie_row.path_changed.connect(self._on_cookie_changed)
        gb1_layout.addWidget(self.cookie_row)

        self.master_row = FileRowWidget("📦 Master (Combo)", "Excel Files (*.xlsx)")
        self.master_row.set_path(self._master_real)
        self.master_row.path_changed.connect(self._on_master_changed)
        gb1_layout.addWidget(self.master_row)

        self.retail_row = FileRowWidget("🛍 Bán lẻ", "Excel Files (*.xlsx)")
        self.retail_row.set_path(self._retail_real)
        self.retail_row.path_changed.connect(self._on_retail_changed)
        gb1_layout.addWidget(self.retail_row)

        self.template_row = FileRowWidget("📋 Mẫu xuất hàng", "Excel Files (*.xlsx)")
        self.template_row.set_path(self._template_real)
        self.template_row.path_changed.connect(self._on_template_changed)
        gb1_layout.addWidget(self.template_row)

        self.output_row = FileRowWidget("📂 Thư mục lưu", is_dir=True)
        self.output_row.set_path(str(BASE_DIR / "outputs"))
        self.output_row._update_status_dir()
        self.output_row.path_changed.connect(self._on_output_dir_changed)
        gb1_layout.addWidget(self.output_row)

        layout.addWidget(gb1)

        # ── Group 2: Xem trước dữ liệu Excel & Ô Tìm Kiếm ──
        gb2 = QGroupBox("Xem trước dữ liệu Excel (Hiển thị tối đa 5000 dòng)")
        gb2_layout = QVBoxLayout(gb2)
        gb2_layout.setSpacing(12)
        gb2_layout.setContentsMargins(0, 0, 0, 0)

        toolbar = QHBoxLayout()
        toolbar.setSpacing(8)

        self.preview_btn_group = QButtonGroup(self)
        self.preview_btn_group.setExclusive(True)

        self.preview_master_btn = QPushButton("Bảng Master (Combo)")
        self.preview_master_btn.setObjectName("previewTabBtn")
        self.preview_master_btn.setCheckable(True)
        self.preview_master_btn.setChecked(True)
        self.preview_master_btn.setProperty("active", True)
        self.preview_master_btn.setCursor(Qt.PointingHandCursor)
        self.preview_btn_group.addButton(self.preview_master_btn, 0)
        toolbar.addWidget(self.preview_master_btn)

        self.preview_retail_btn = QPushButton("Bảng Bán lẻ")
        self.preview_retail_btn.setObjectName("previewTabBtn")
        self.preview_retail_btn.setCheckable(True)
        self.preview_retail_btn.setCursor(Qt.PointingHandCursor)
        self.preview_btn_group.addButton(self.preview_retail_btn, 1)
        toolbar.addWidget(self.preview_retail_btn)
        self.preview_btn_group.idClicked.connect(self._on_preview_tab_changed)

        toolbar.addStretch()

        self.add_row_btn = QPushButton("➕ Thêm dòng")
        self.add_row_btn.setObjectName("smallBtn")
        self.add_row_btn.setCursor(Qt.PointingHandCursor)
        self.add_row_btn.clicked.connect(self._on_add_row)
        toolbar.addWidget(self.add_row_btn)

        self.del_row_btn = QPushButton("🗑 Xóa dòng")
        self.del_row_btn.setObjectName("smallBtn")
        self.del_row_btn.setCursor(Qt.PointingHandCursor)
        self.del_row_btn.clicked.connect(self._on_delete_rows)
        toolbar.addWidget(self.del_row_btn)

        self.search_input = QLineEdit()
        self.search_input.setObjectName("searchBox")
        self.search_input.setPlaceholderText("🔍 Tìm kiếm SKU, tên sản phẩm...")
        self.search_input.setFixedWidth(250)
        self.search_input.textChanged.connect(self._filter_preview_table)
        toolbar.addWidget(self.search_input)

        gb2_layout.addLayout(toolbar)

        self.master_table = QTableWidget()
        self.master_table.setEditTriggers(QTableWidget.DoubleClicked)
        self.master_table.setMinimumHeight(450) # Đảm bảo bảng luôn mở rộng thoải mái để xem
        self.master_table.cellChanged.connect(self._on_table_cell_changed)
        gb2_layout.addWidget(self.master_table, 1)

        self.retail_table = QTableWidget()
        self.retail_table.setEditTriggers(QTableWidget.DoubleClicked)
        self.retail_table.setMinimumHeight(450) # Đảm bảo bảng luôn mở rộng thoải mái để xem
        self.retail_table.cellChanged.connect(self._on_table_cell_changed)
        self.retail_table.hide()
        gb2_layout.addWidget(self.retail_table, 1)

        layout.addWidget(gb2, 1)
        scroll.setWidget(w)
        return scroll

    # ═══════════════════════════════════════════════════════
    # TAB 1: CẤU HÌNH IN 
    # ═══════════════════════════════════════════════════════
    def _build_print_tab(self):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)

        w = QWidget()
        w.setObjectName("scrollContent")
        layout = QVBoxLayout(w)
        layout.setContentsMargins(24, 24, 24, 24)

        gb = QGroupBox("Thiết lập tải và in tự động")
        gb_layout = QVBoxLayout(gb)
        gb_layout.setSpacing(12)
        gb_layout.setContentsMargins(0, 0, 0, 0)

        # ── Carrier grid: 2 cột ──
        grid = QGridLayout()
        grid.setHorizontalSpacing(32)
        grid.setVerticalSpacing(16)

        self.carrier_spinboxes = {}
        self.carrier_checkboxes = {}
        carriers = [
            ("SPX Express:", "spx"), ("Giao Hàng Nhanh:", "ghn"), ("Ninja Van:", "nv"),
            ("VNPost Nhanh:", "vnp"), ("J&T Express:", "jt"), ("BEST Express:", "best"),
            ("Viettel Post:", "vt"), ("VNP - Hàng Cồng Kềnh:", "vnp_bulky"),
            ("GHN - Hàng Cồng Kềnh:", "ghn_bulky"), ("NJV - Hàng Cồng Kềnh:", "njv_bulky"),
            ("SPX - Hàng Cồng Kềnh:", "spx_bulky"), ("VTP - Hàng Cồng Kềnh:", "vtp_bulky"),
            ("Ahamove - Trong Ngày:", "ahamove"), ("Ahamove SBS:", "ahamove_sbs"),
            ("SPX Express - TN:", "spx_tn"), ("SPX Express SBS:", "spx_sbs"),
            ("Tủ nhận hàng - SPX:", "spx_locker"), ("ĐVVC khác:", "other"),
        ]
        for idx, (label, key) in enumerate(carriers):
            row, col = idx % 3, (idx // 3)

            box = QWidget()
            box_ly = QHBoxLayout(box)
            box_ly.setContentsMargins(0,0,0,0)

            cb = QCheckBox(label.replace(":", ""))
            cb.setChecked(True)  # mặc định bật tất cả
            cb.setStyleSheet("font-weight: 500;")
            cb.setFixedWidth(120)
            box_ly.addWidget(cb)
            self.carrier_checkboxes[key] = cb

            sb = QSpinBox()
            sb.setRange(0, 9999)
            sb.setValue(0)
            sb.setFixedWidth(65)
            sb.setAlignment(Qt.AlignRight)
            sb.setToolTip("0 = tất cả đơn của hãng này")
            box_ly.addWidget(sb)

            hint = QLabel("đơn  (0 = tất cả)")
            hint.setStyleSheet("color: #64748B; font-size: 11px;")
            box_ly.addWidget(hint)

            # Checkbox OFF → disable spinbox (bỏ qua hãng này)
            cb.toggled.connect(sb.setEnabled)

            box_ly.addStretch()

            grid.addWidget(box, row, col)
            self.carrier_spinboxes[key] = sb

        gb_layout.addLayout(grid)

        # Divider
        div = QFrame()
        div.setObjectName("divider")
        div.setFrameShape(QFrame.HLine)
        gb_layout.addWidget(div)

        # ── In tự động row ──
        print_row = QHBoxLayout()
        print_row.setSpacing(12)

        self.auto_print_cb = QCheckBox("🖨️ In tự động ra máy in")
        self.auto_print_cb.setChecked(True)
        self.auto_print_cb.setStyleSheet("font-weight: bold; color: #059669;")
        print_row.addWidget(self.auto_print_cb)

        printers_list = _get_printers()
        self.printer_combo = QComboBox()
        self.printer_combo.addItems(printers_list)
        self.printer_combo.setMinimumWidth(180)
        if printers_list:
            default_printer = _get_default_printer()
            if default_printer and default_printer in printers_list:
                self.printer_combo.setCurrentIndex(printers_list.index(default_printer))
            else:
                self.printer_combo.setCurrentIndex(0)
        print_row.addWidget(self.printer_combo, 1)

        refresh_btn = QPushButton("↻ Cập nhật")
        refresh_btn.setObjectName("refreshBtn")
        refresh_btn.setCursor(Qt.PointingHandCursor)
        refresh_btn.clicked.connect(self._refresh_printers)
        print_row.addWidget(refresh_btn)

        gb_layout.addLayout(print_row)

        # ── Paper settings row ──
        paper_row = QHBoxLayout()
        paper_row.setSpacing(12)
        paper_row.setContentsMargins(0, 8, 0, 0)

        paper_row.addWidget(QLabel("In mặt:"))
        self.duplex_combo = QComboBox()
        self.duplex_combo.addItems(["simplex", "longedge", "shortedge"])
        self.duplex_combo.setFixedWidth(100)
        paper_row.addWidget(self.duplex_combo)

        paper_row.addWidget(QLabel("  Batch in (tờ/lần):"))
        self.batch_size_spin = QSpinBox()
        self.batch_size_spin.setRange(0, 100)
        self.batch_size_spin.setValue(0)
        self.batch_size_spin.setFixedWidth(60)
        self.batch_size_spin.setToolTip("0 = không chia batch. Nhập số >0 để chia nhỏ file in")
        self.batch_size_spin.setSpecialValueText("0 (không chia)")
        paper_row.addWidget(self.batch_size_spin)

        paper_row.addStretch()
        gb_layout.addLayout(paper_row)

        # ── Engine in PDF: Foxit PDF Reader (XPS Print Path, spool ~15MB) ──
        foxit_detected = _find_foxit_exe()

        # Label hiển thị engine in PDF
        engine_label = QLabel("Engine in PDF:")
        engine_label.setFixedWidth(130)
        engine_label.setStyleSheet("font-weight: 600; color: #1E293B; font-size: 10pt;")

        foxit_status = QLabel()
        foxit_status.setWordWrap(True)
        if foxit_detected:
            foxit_status.setText("✅ Foxit PDF Reader — XPS Print Path (spool ~15MB)")
            foxit_status.setStyleSheet("color: #059669; font-size: 9pt; padding: 2px 0;")
        else:
            foxit_status.setText("⚠ Chưa cài Foxit PDF Reader — Vui lòng cài để in file PDF!")
            foxit_status.setStyleSheet("color: #DC2626; font-weight: 600; font-size: 10pt; padding: 2px 0;")

        gb_layout.addWidget(engine_label)
        gb_layout.addWidget(foxit_status)

        # ── Chrome info ──
        chrome_info = QLabel("🌐 Sử dụng Google Chrome có sẵn trên máy")
        chrome_info.setStyleSheet("color: #059669; font-weight: 500; font-size: 13px; margin-top: 8px;")
        gb_layout.addWidget(chrome_info)

        # ── Loại trừ đơn bán trước ──
        self.exclude_pre_orders_cb = QCheckBox("🚫 Loại trừ đơn bán trước (Pre-order) khi tải đơn")
        self.exclude_pre_orders_cb.setChecked(False)
        self.exclude_pre_orders_cb.setToolTip("Bỏ tick nếu bạn MUỐN in cả đơn bán trước.\nTick để bỏ qua đơn bán trước, chỉ in đơn thường.")
        self.exclude_pre_orders_cb.setStyleSheet("color: #64748B; font-weight: 500; font-size: 13px; margin-top: 8px;")
        gb_layout.addWidget(self.exclude_pre_orders_cb)

        # ── Test mode ──
        self.test_mode_cb = QCheckBox("🧪 Bật chế độ Test Mode (Chỉ tải danh sách đơn, KHÔNG thao tác in)")
        self.test_mode_cb.setChecked(False)
        self.test_mode_cb.setStyleSheet("color: #D97706; font-weight: 600; font-size: 14px; margin-top: 12px;")
        gb_layout.addWidget(self.test_mode_cb)

        layout.addWidget(gb)
        layout.addStretch()
        scroll.setWidget(w)
        return scroll

    # ═══════════════════════════════════════════════════════
    # TAB 2: LỊCH TRÌNH
    # ═══════════════════════════════════════════════════════
    def _build_schedule_tab(self):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)

        w = QWidget()
        w.setObjectName("scrollContent")
        layout = QVBoxLayout(w)
        layout.setContentsMargins(24, 24, 24, 24)

        gb = QGroupBox("Thiết lập bộ đếm thời gian")
        gb_layout = QVBoxLayout(gb)
        gb_layout.setSpacing(16)
        gb_layout.setContentsMargins(0, 0, 0, 0)

        self.sched_button_group = QButtonGroup(self)

        self.once_rb = QRadioButton("▶ Chạy ngay lập tức 1 lần duy nhất")
        self.once_rb.setChecked(False)
        self.once_rb.setProperty("mode", "once")
        self.sched_button_group.addButton(self.once_rb)
        gb_layout.addWidget(self.once_rb)

        self.interval_rb = QRadioButton("🔄 Lặp lại tự động mỗi N giờ")
        self.interval_rb.setProperty("mode", "interval")
        self.sched_button_group.addButton(self.interval_rb)
        gb_layout.addWidget(self.interval_rb)

        # Interval sub-panel
        self.interval_panel = QWidget()
        ip_layout = QHBoxLayout(self.interval_panel)
        ip_layout.setContentsMargins(32, 0, 0, 0)
        ip_layout.setSpacing(8)
        ip_layout.addWidget(QLabel("Khởi chạy lại mỗi:"))
        self.interval_spin = QSpinBox()
        self.interval_spin.setRange(1, 24)
        self.interval_spin.setValue(1)
        self.interval_spin.setFixedWidth(80)
        ip_layout.addWidget(self.interval_spin)
        lbl_h = QLabel("giờ")
        lbl_h.setStyleSheet("color: #64748B;")
        ip_layout.addWidget(lbl_h)
        ip_layout.addStretch()
        self.interval_panel.hide()
        gb_layout.addWidget(self.interval_panel)

        self.weekly_rb = QRadioButton("📅 Chạy theo lịch hàng tuần")
        self.weekly_rb.setChecked(True)
        self.weekly_rb.setProperty("mode", "weekly")
        self.sched_button_group.addButton(self.weekly_rb)
        gb_layout.addWidget(self.weekly_rb)

        # Weekly sub-panel — mỗi ngày trong tuần có checkbox + khung giờ riêng
        self.weekly_panel = QWidget()
        wp_outer = QVBoxLayout(self.weekly_panel)
        wp_outer.setContentsMargins(32, 0, 0, 0)
        wp_outer.setSpacing(8)

        # Quick-fill row
        quick_row = QHBoxLayout()
        quick_row.setSpacing(8)
        quick_row.addWidget(QLabel("Nhập giờ mẫu:"))
        self.weekly_master_time_edit = QLineEdit("07:00, 13:00, 15:00, 16:00, 17:00, 18:00")
        self.weekly_master_time_edit.setFixedWidth(200)
        self.weekly_master_time_edit.setToolTip("Định dạng HH:MM, phân cách bằng dấu phẩy")
        quick_row.addWidget(self.weekly_master_time_edit)
        quick_row.addWidget(QLabel("(phân cách bằng dấu phẩy)"))
        apply_btn = QPushButton("Áp dụng cho tất cả")
        apply_btn.setFixedWidth(160)
        apply_btn.setCursor(Qt.PointingHandCursor)
        apply_btn.clicked.connect(self._on_weekly_apply_all)
        quick_row.addWidget(apply_btn)
        quick_row.addStretch()
        wp_outer.addLayout(quick_row)

        # Separator
        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet("color: #E2E8F0; max-height: 1px;")
        wp_outer.addWidget(sep)

        # 7 day rows in a vertical layout
        day_names = ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7", "Chủ Nhật"]
        for idx, name in enumerate(day_names):
            day_row = QHBoxLayout()
            day_row.setSpacing(8)

            cb = QCheckBox(name)
            cb.setFixedWidth(90)
            cb.setChecked(idx < 6)  # Mặc định T2-T7 checked, CN unchecked
            cb.setStyleSheet("font-weight: 500;")
            self.weekly_day_checkboxes[idx] = cb

            te = QLineEdit("07:00, 13:00, 15:00, 16:00, 17:00, 18:00" if idx < 6 else "")
            te.setFixedWidth(220)
            te.setEnabled(idx < 6)
            te.setPlaceholderText("VD: 08:00, 14:00")
            self.weekly_day_time_edits[idx] = te

            # Checkbox toggle -> enable/disable time edit
            cb.toggled.connect(lambda checked, i=idx: self.weekly_day_time_edits[i].setEnabled(checked))

            day_row.addWidget(cb)
            day_row.addWidget(te)
            day_row.addStretch()
            wp_outer.addLayout(day_row)

        wp_outer.addStretch()
        self.weekly_panel.show()
        gb_layout.addWidget(self.weekly_panel)

        self.sched_button_group.buttonClicked.connect(self._on_schedule_mode_changed)

        layout.addWidget(gb)
        layout.addStretch()
        scroll.setWidget(w)
        return scroll

    # ═══════════════════════════════════════════════════════
    # TAB 3: NHẬT KÝ & KẾT QUẢ
    # ═══════════════════════════════════════════════════════
    def _build_log_tab(self):
        w = QWidget()
        w.setObjectName("scrollContent")
        layout = QVBoxLayout(w)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(24)

        # ── Log group (flex grow) ──
        log_gb = QGroupBox("Nhật ký hệ thống (Log)")
        log_layout = QVBoxLayout(log_gb)
        log_layout.setSpacing(4)
        log_layout.setContentsMargins(0, 0, 0, 0)

        clear_row = QHBoxLayout()
        clear_row.addStretch()
        clear_btn = QPushButton("🗑 Xóa log")
        clear_btn.setObjectName("smallBtn")
        clear_btn.setCursor(Qt.PointingHandCursor)
        clear_btn.clicked.connect(self._clear_log)
        clear_row.addWidget(clear_btn)
        log_layout.addLayout(clear_row)

        self.log_view = QTextEdit()
        self.log_view.setObjectName("logView")
        self.log_view.setReadOnly(True)
        log_layout.addWidget(self.log_view, 1)

        layout.addWidget(log_gb, 1)

        # ── Results group ──
        res_gb = QGroupBox("📁 File kết quả (Click đúp để mở)")
        res_layout = QVBoxLayout(res_gb)
        res_layout.setSpacing(8)
        res_layout.setContentsMargins(0, 0, 0, 0)

        self.result_list = QListWidget()
        self.result_list.setObjectName("resultList")
        self.result_list.setMaximumHeight(100)
        self.result_list.itemDoubleClicked.connect(self._on_open_result)
        res_layout.addWidget(self.result_list)

        res_btns = QHBoxLayout()
        res_btns.addStretch()
        open_folder_btn = QPushButton("📂 Mở thư mục")
        open_folder_btn.setObjectName("smallBtn")
        open_folder_btn.setCursor(Qt.PointingHandCursor)
        open_folder_btn.clicked.connect(self._open_output_dir)
        res_btns.addWidget(open_folder_btn)
        
        clear_res_btn = QPushButton("🗑 Xóa danh sách")
        clear_res_btn.setObjectName("smallBtn")
        clear_res_btn.setCursor(Qt.PointingHandCursor)
        clear_res_btn.clicked.connect(self._clear_results)
        res_btns.addWidget(clear_res_btn)
        
        res_layout.addLayout(res_btns)

        layout.addWidget(res_gb)
        return w

    # ═══════════════════════════════════════════════════════
    # TAB 4: TEST THỦ CÔNG
    # ═══════════════════════════════════════════════════════
    def _build_test_tab(self):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)

        w = QWidget()
        w.setObjectName("scrollContent")
        layout = QVBoxLayout(w)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(16)

        # ── Group 1: Tính toán từ Picking list ──
        gb1 = QGroupBox("📊 Tính toán từ Picking list có sẵn")
        gb1_layout = QVBoxLayout(gb1)
        gb1_layout.setSpacing(8)

        btn_row1 = QHBoxLayout()
        btn_pick = QPushButton("📂 Chọn file Picking list...")
        btn_pick.setObjectName("browseBtn")
        btn_pick.setCursor(Qt.PointingHandCursor)
        btn_pick.clicked.connect(self._test_select_picking)
        btn_row1.addWidget(btn_pick)

        btn_clear1 = QPushButton("🗑 Xóa danh sách")
        btn_clear1.setObjectName("smallBtn")
        btn_clear1.setCursor(Qt.PointingHandCursor)
        btn_clear1.clicked.connect(lambda: self._test_picking_list.clear())
        btn_row1.addWidget(btn_clear1)
        btn_row1.addStretch()
        gb1_layout.addLayout(btn_row1)

        self._test_picking_list = QListWidget()
        self._test_picking_list.setMaximumHeight(100)
        self._test_picking_list.setObjectName("resultList")
        gb1_layout.addWidget(self._test_picking_list)

        btn_calc = QPushButton("▶ Chạy tính toán")
        btn_calc.setObjectName("schedBtn")
        btn_calc.setCursor(Qt.PointingHandCursor)
        btn_calc.clicked.connect(self._test_run_calculator)
        gb1_layout.addWidget(btn_calc)

        layout.addWidget(gb1)

        # ── Group 2: In shipping label ──
        gb2 = QGroupBox("🖨️ In Shipping label có sẵn")
        gb2_layout = QVBoxLayout(gb2)
        gb2_layout.setSpacing(8)

        btn_row2 = QHBoxLayout()
        btn_ship = QPushButton("📂 Chọn file Shipping label...")
        btn_ship.setObjectName("browseBtn")
        btn_ship.setCursor(Qt.PointingHandCursor)
        btn_ship.clicked.connect(self._test_select_shipping)
        btn_row2.addWidget(btn_ship)

        btn_clear2 = QPushButton("🗑 Xóa danh sách")
        btn_clear2.setObjectName("smallBtn")
        btn_clear2.setCursor(Qt.PointingHandCursor)
        btn_clear2.clicked.connect(lambda: self._test_shipping_list.clear())
        btn_row2.addWidget(btn_clear2)
        btn_row2.addStretch()
        gb2_layout.addLayout(btn_row2)

        self._test_shipping_list = QListWidget()
        self._test_shipping_list.setMaximumHeight(100)
        self._test_shipping_list.setObjectName("resultList")
        gb2_layout.addWidget(self._test_shipping_list)

        btn_print = QPushButton("🖨️ In + Merge 2-up")
        btn_print.setObjectName("schedBtn")
        btn_print.setCursor(Qt.PointingHandCursor)
        btn_print.clicked.connect(self._test_print_shipping)
        gb2_layout.addWidget(btn_print)

        layout.addWidget(gb2)

        # ── Group 3: Tự động toàn bộ (tính toán + in) ──
        gb3 = QGroupBox("🚀 Tự động toàn bộ (chọn tất cả file đã tải về)")
        gb3_layout = QVBoxLayout(gb3)
        gb3_layout.setSpacing(8)

        btn_auto = QPushButton("📂 Chọn tất cả file của 1 lần tải...")
        btn_auto.setObjectName("browseBtn")
        btn_auto.setCursor(Qt.PointingHandCursor)
        btn_auto.clicked.connect(self._test_select_all)
        gb3_layout.addWidget(btn_auto)

        self._test_all_list = QListWidget()
        self._test_all_list.setMaximumHeight(100)
        self._test_all_list.setObjectName("resultList")
        gb3_layout.addWidget(self._test_all_list)

        btn_row3 = QHBoxLayout()
        btn_run_all = QPushButton("▶ Chạy toàn bộ (Tính toán + In)")
        btn_run_all.setObjectName("schedBtn")
        btn_run_all.setCursor(Qt.PointingHandCursor)
        btn_run_all.clicked.connect(self._test_run_all)
        btn_row3.addWidget(btn_run_all, 2)

        btn_clear3 = QPushButton("🗑 Xóa danh sách")
        btn_clear3.setObjectName("smallBtn")
        btn_clear3.setCursor(Qt.PointingHandCursor)
        btn_clear3.clicked.connect(lambda: self._test_all_list.clear())
        btn_row3.addWidget(btn_clear3)
        gb3_layout.addLayout(btn_row3)

        layout.addWidget(gb3)
        layout.addStretch()
        scroll.setWidget(w)
        return scroll

    def _test_select_picking(self):
        files, _ = QFileDialog.getOpenFileNames(self, "Chọn file Picking list", "",
                                                 "PDF Files (*.pdf)")
        for f in files:
            self._test_picking_list.addItem(f)

    def _test_select_shipping(self):
        files, _ = QFileDialog.getOpenFileNames(self, "Chọn file Shipping label", "",
                                                 "PDF Files (*.pdf)")
        for f in files:
            self._test_shipping_list.addItem(f)

    def _test_run_calculator(self):
        pdfs = [self._test_picking_list.item(i).text()
                for i in range(self._test_picking_list.count())]
        if not pdfs:
            QMessageBox.warning(self, "Cảnh báo", "Chọn ít nhất 1 file Picking list.")
            return
        master = self._master_real
        retail = self._retail_real
        template = self._template_real
        if not master or not Path(master).exists():
            QMessageBox.critical(self, "Lỗi", "Chọn file Master (Combo) hợp lệ ở tab Tệp dữ liệu.")
            return
        if not template or not Path(template).exists():
            QMessageBox.critical(self, "Lỗi", "Chọn file Mẫu xuất hàng hợp lệ ở tab Tệp dữ liệu.")
            return
        out_dir = self.output_row.get_real_path() or str(BASE_DIR / "outputs")

        def _run():
            self._test_log.emit("info", "🧪 TEST: Bắt đầu tính toán...")
            # Nếu file gộp → process_all() sẽ tự tách phiếu xuất để tính
            try:
                results = run_calculator(pdfs, out_dir, master, retail, template,
                                         lambda m, t='': self._test_log.emit(t, m))
                for r in results:
                    self._test_log.emit("ok", f"  ✓ {r['rows']} SKU | Qty={r['tong_qty']} | Sold={r['tong_sold']} | Promo={r['tong_promo']}")
                    for key, lbl in [('xlsx_report', '📊')]:
                        fp = r['files'].get(key)
                        if fp and Path(fp).exists():
                            self._add_result(fp, label=lbl)
                self._test_log.emit("bold_ok", "✅ Tính toán hoàn tất")
            except Exception as e:
                self._test_log.emit("err", f"✗ Lỗi: {e}")

        threading.Thread(target=_run, daemon=True).start()

    def _test_print_shipping(self):
        files = [self._test_shipping_list.item(i).text()
                 for i in range(self._test_shipping_list.count())]
        if not files:
            QMessageBox.warning(self, "Cảnh báo", "Chọn ít nhất 1 file Shipping label.")
            return
        pdf_settings = self._build_pdf_settings()
        batch_size = self.batch_size_spin.value()
        do_print = self.auto_print_cb.isChecked()
        printer = self.printer_combo.currentText()
        if do_print and not printer:
            QMessageBox.warning(self, "Cảnh báo", "Chọn máy in ở tab Cấu hình In.")
            return

        def _run():
            self._test_log.emit("info", f"🧪 TEST: Xử lý {len(files)} file...")
            for fp in files:
                try:
                    # Nếu là file gộp → tự tách lấy shipping label
                    actual_fp = fp
                    if _get_calculator()[3](fp) == 'shopee':
                        from calculator import split_shopee_pdf
                        out_dir = os.path.dirname(fp) or 'outputs'
                        _, shipping_path, _ = split_shopee_pdf(fp, out_dir)
                        if shipping_path:
                            self._test_log.emit("info", f"  📦 Đã tách shipping label: {Path(shipping_path).name}")
                            actual_fp = shipping_path

                    if do_print:
                        # Luôn merge 2-up trước khi in shipping label
                        merged = _merge_pdf_2up(actual_fp)
                        print_target = merged if merged else actual_fp
                        if merged:
                            self._test_log.emit("info", f"  📄 Đã merge 2-up: {Path(merged).name}")
                        _print_file(print_target, printer, pdf_settings=pdf_settings, batch_size=batch_size,
                                    log_cb=lambda m, t='': self._test_log.emit(t, m))
                        self._test_log.emit("ok", f"  ✓ Đã gửi in: {Path(actual_fp).name}")
                    else:
                        # Chỉ merge 2-up, không in (tick 'In tự động' để in)
                        merged = _merge_pdf_2up(actual_fp)
                        if merged:
                            self._add_result(merged, label='📄')
                            self._test_log.emit("ok", f"  ✓ Đã merge 2-up: {Path(merged).name}")
                            self._test_log.emit("dim", "  ℹ Chưa in — tick 'In tự động ra máy in' ở tab Cấu hình In để in")
                        else:
                            self._test_log.emit("ok", f"  ✓ File sẵn sàng: {Path(actual_fp).name}")
                    # ── Dọn file shipping label trung gian đã tách ──
                    if actual_fp != fp and Path(actual_fp).exists():
                        try:
                            Path(actual_fp).unlink()
                            self._test_log.emit("dim", f"  🗑 Đã dọn: {Path(actual_fp).name}")
                        except Exception:
                            pass
                except Exception as e:
                    self._test_log.emit("err", f"  ✗ Lỗi xử lý {Path(fp).name}: {e}")
            self._test_log.emit("bold_ok", "✅ Hoàn tất")

        threading.Thread(target=_run, daemon=True).start()

    def _test_select_all(self):
        """Chọn tất cả file PDF đã tải về (cả Picking list + Shipping label)."""
        files, _ = QFileDialog.getOpenFileNames(self, "Chọn tất cả file đã tải về", "",
                                                 "PDF Files (*.pdf)")
        for f in files:
            self._test_all_list.addItem(f)

    def _test_run_all(self):
        """Tự động: chọn file Shopee PDF gộp → tách → tính toán → in."""
        all_files = [self._test_all_list.item(i).text()
                     for i in range(self._test_all_list.count())]
        if not all_files:
            QMessageBox.warning(self, "Cảnh báo", "Chọn ít nhất 1 file PDF.")
            return

        master = self._master_real
        retail = self._retail_real
        out_dir = self.output_row.get_real_path() or str(BASE_DIR / "outputs")
        printer = self.printer_combo.currentText()
        auto_print = self.auto_print_cb.isChecked()
        pdf_settings = self._build_pdf_settings()
        batch_size = self.batch_size_spin.value()

        def _run():
            do_print = auto_print and printer
            if not do_print:
                self._test_log.emit("warn", "⚠ In bị tắt — tick 'In tự động ra máy in' ở tab Cấu hình In để in")

            # ── Bước 1: Tự động tách file gộp Shopee ──
            picking_files = []
            shipping_files = []

            for f in all_files:
                fname = Path(f).name.lower()
                if _get_calculator()[3](f) == 'shopee':
                    # Tự tách file gộp
                    from calculator import split_shopee_pdf
                    out_dir = os.path.dirname(f) or 'outputs'
                    picking_path, shipping_path, _ = split_shopee_pdf(f, out_dir)
                    if picking_path:
                        self._test_log.emit("info", f"  📋 Tách phiếu xuất: {Path(picking_path).name}")
                        picking_files.append(picking_path)
                    if shipping_path:
                        self._test_log.emit("info", f"  📦 Tách shipping label: {Path(shipping_path).name}")
                        shipping_files.append(shipping_path)
                    if not picking_path and not shipping_path:
                        # Không tách được → cho vào picking để thử tính
                        picking_files.append(f)
                elif '_shipping_label' in fname or 'shipping' in fname:
                    shipping_files.append(f)
                else:
                    picking_files.append(f)

            # ── Bước 2: In shipping label (merge 2-up) ──
            if do_print and shipping_files:
                self._test_log.emit("info", f"🖨️ In {len(shipping_files)} file shipping label...")
                for fp in shipping_files:
                    try:
                        _print_file(fp, printer, pdf_settings=pdf_settings, batch_size=batch_size,
                                    log_cb=lambda m, t='': self._test_log.emit(t, m))
                        self._test_log.emit("ok", f"  🖨️ {Path(fp).name}")
                    except Exception as e:
                        self._test_log.emit("err", f"  ✗ Lỗi in {Path(fp).name}: {e}")

            # ── Dọn file shipping label trung gian đã tách ──
            for fp in shipping_files:
                if fp not in all_files and Path(fp).exists():
                    try:
                        Path(fp).unlink()
                        self._test_log.emit("dim", f"  🗑 Đã dọn: {Path(fp).name}")
                    except Exception:
                        pass

            # ── Bước 3: Tính toán + in báo cáo ──
            if picking_files:
                self._test_log.emit("info", f"📊 {len(picking_files)} file → tính toán...")
                try:
                    results = run_calculator(picking_files, out_dir, master, retail, self._template_real,
                                             lambda m, t='': self._test_log.emit(t, m))
                    for r in results:
                        self._test_log.emit("ok", f"  ✓ {r['rows']} SKU | Qty={r['tong_qty']} | Sold={r['tong_sold']} | Promo={r['tong_promo']}")
                        # Thêm file kết quả vào list
                        for key in ('xlsx_report', 'pdf_report'):
                            fp = r['files'].get(key)
                            if fp and Path(fp).exists():
                                self._add_result(fp)
                                if do_print and key == 'pdf_report':
                                    try:
                                        for copy_num in [1, 2]:
                                            self._test_log.emit("info", f"  🖨️ In bản {copy_num}/2: {Path(fp).name}")
                                            _print_file(fp, printer, pdf_settings=pdf_settings, batch_size=batch_size,
                                                        log_cb=lambda m, t='': self._test_log.emit(t, m))
                                            if copy_num == 1:
                                                import time as _t3; _t3.sleep(2)
                                        self._test_log.emit("ok", f"  🖨️ Báo cáo 2 bản: {Path(fp).name}")
                                    except Exception as e:
                                        self._test_log.emit("err", f"  ✗ Lỗi in báo cáo: {e}")
                except Exception as e:
                    self._test_log.emit("err", f"  ✗ Lỗi tính toán: {e}")

            # ── Dọn file phiếu xuất trung gian đã tách ──
            for fp in picking_files:
                if fp not in all_files and Path(fp).exists():
                    try:
                        Path(fp).unlink()
                        self._test_log.emit("dim", f"  🗑 Đã dọn: {Path(fp).name}")
                    except Exception:
                        pass

            self._test_log.emit("bold_ok", "✅ Hoàn tất toàn bộ")

        threading.Thread(target=_run, daemon=True).start()

    # ═══════════════════════════════════════════════════════
    # TAB 5: TỔNG HỢP BÁO CÁO
    # ═══════════════════════════════════════════════════════
    def _build_aggregate_tab(self):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)

        w = QWidget()
        w.setObjectName("scrollContent")
        layout = QVBoxLayout(w)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(16)

        gb = QGroupBox("📊 Tổng hợp nhiều file báo cáo thành 1 file duy nhất")
        gb_layout = QVBoxLayout(gb)
        gb_layout.setSpacing(8)

        # ── Description ──
        desc = QLabel("Chọn thư mục chứa các file Excel báo cáo (Phieu_xuat_hang_*.xlsx).\n"
                      "Dữ liệu sẽ được cộng dồn theo SKU từ tất cả các file trong thư mục.")
        desc.setStyleSheet("color: #64748B; font-size: 12px; padding: 4px 0;")
        desc.setWordWrap(True)
        gb_layout.addWidget(desc)

        # ── Quick: Tổng hợp hôm nay ──
        quick_row = QHBoxLayout()
        self._aggregate_today_btn = QPushButton("📅 Tổng hợp hôm nay")
        self._aggregate_today_btn.setObjectName("schedBtn")
        self._aggregate_today_btn.setCursor(Qt.PointingHandCursor)
        self._aggregate_today_btn.clicked.connect(self._on_aggregate_today)
        quick_row.addWidget(self._aggregate_today_btn)
        quick_row.addStretch()
        gb_layout.addLayout(quick_row)

        # ── Separator ──
        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet("color: #E2E8F0; max-height: 1px;")
        gb_layout.addWidget(sep)

        # ── Folder selection ──
        btn_row = QHBoxLayout()
        btn_add = QPushButton("📂 Chọn thư mục...")
        btn_add.setObjectName("browseBtn")
        btn_add.setCursor(Qt.PointingHandCursor)
        btn_add.clicked.connect(self._aggregate_select_folder)
        btn_row.addWidget(btn_add)

        btn_clear = QPushButton("🗑 Xóa danh sách")
        btn_clear.setObjectName("smallBtn")
        btn_clear.setCursor(Qt.PointingHandCursor)
        btn_clear.clicked.connect(lambda: self._aggregate_file_list.clear())
        btn_row.addWidget(btn_clear)
        btn_row.addStretch()
        gb_layout.addLayout(btn_row)

        self._aggregate_folder_label = QLabel("")
        self._aggregate_folder_label.setStyleSheet("color: #60A5FA; font-size: 11px;")
        gb_layout.addWidget(self._aggregate_folder_label)

        self._aggregate_file_list = QListWidget()
        self._aggregate_file_list.setMaximumHeight(150)
        self._aggregate_file_list.setObjectName("resultList")
        gb_layout.addWidget(self._aggregate_file_list)

        # ── Run button ──
        self._aggregate_run_btn = QPushButton("▶ Tổng hợp")
        self._aggregate_run_btn.setObjectName("schedBtn")
        self._aggregate_run_btn.setCursor(Qt.PointingHandCursor)
        self._aggregate_run_btn.clicked.connect(self._on_run_aggregate)
        gb_layout.addWidget(self._aggregate_run_btn)

        # ── Aggregate log ──
        self._aggregate_log = QTextEdit()
        self._aggregate_log.setObjectName("logView")
        self._aggregate_log.setReadOnly(True)
        self._aggregate_log.setMaximumHeight(250)
        gb_layout.addWidget(self._aggregate_log)

        layout.addWidget(gb)
        layout.addStretch()
        scroll.setWidget(w)
        return scroll

    def _aggregate_select_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Chọn thư mục chứa file báo cáo", "")
        if not folder:
            return
        self._aggregate_file_list.clear()
        self._aggregate_folder_label.setText(f"📁 {folder}")
        # Auto-scan folder for report files
        import glob as _glob
        patterns = [
            str(Path(folder) / 'Phieu_xuat_hang_*.xlsx'),
            str(Path(folder) / '*.xlsx'),
        ]
        found = set()
        for pat in patterns:
            for f in _glob.glob(pat):
                found.add(f)
        for f in sorted(found):
            self._aggregate_file_list.addItem(f)
        if not found:
            self._ag_log("warn", f"Không tìm thấy file Excel nào trong thư mục")

    def _on_run_aggregate(self):
        files = [self._aggregate_file_list.item(i).text()
                 for i in range(self._aggregate_file_list.count())]
        if not files:
            QMessageBox.warning(self, "Cảnh báo", "Chọn ít nhất 1 file báo cáo Excel.")
            return
        template = self._template_real
        if not template or not Path(template).exists():
            QMessageBox.critical(self, "Lỗi", "Chọn file Mẫu xuất hàng hợp lệ ở tab Tệp dữ liệu.")
            return
        out_dir = self.output_row.get_real_path() or str(BASE_DIR / "outputs")

        self._aggregate_log.clear()
        self._aggregate_run_btn.setEnabled(False)

        def _run():
            self._ag_log("info", f"📊 Bắt đầu tổng hợp {len(files)} file...")
            for f in files:
                self._ag_log("dim", f"  📄 {Path(f).name}")
            try:
                result = _get_calculator()[2](files, out_dir, template)
                self._ag_log("ok", f"  ✓ {result['rows']} SKU | Qty={result['tong_qty']} | "
                                    f"Sold={result['tong_sold']} | Promo={result['tong_promo']}")
                fp = result['files'].get('xlsx_report')
                if fp and Path(fp).exists():
                    self._add_result(fp)
                    self._ag_log("ok", f"  📊 {Path(fp).name}")
                pdf_fp = result['files'].get('pdf_report')
                if pdf_fp and Path(pdf_fp).exists():
                    self._add_result(pdf_fp, label='📄')
                self._ag_log("bold_ok", "✅ Tổng hợp hoàn tất")
            except Exception as e:
                self._ag_log("err", f"✗ Lỗi: {e}")
            finally:
                self._aggregate_run_btn.setEnabled(True)

        threading.Thread(target=_run, daemon=True).start()

    def _ag_log(self, tag: str, msg: str):
        """Ghi log vào aggregate log view (thread-safe qua signal)."""
        self._test_log.emit(tag, f"[Tổng hợp] {msg}")

    def _on_aggregate_today(self):
        """Quét thư mục hôm nay và tổng hợp tất cả báo cáo."""
        base = self.output_row.get_real_path() or str(BASE_DIR / "outputs")
        today_str = datetime.now().strftime('%Y-%m-%d')
        today_dir = Path(base) / today_str
        if not today_dir.exists():
            QMessageBox.warning(self, "Cảnh báo", f"Thư mục hôm nay chưa tồn tại:\n{today_dir}")
            return

        template = self._template_real
        if not template or not Path(template).exists():
            QMessageBox.critical(self, "Lỗi", "Chọn file Mẫu xuất hàng hợp lệ ở tab Tệp dữ liệu.")
            return

        self._aggregate_log.clear()
        self._aggregate_today_btn.setEnabled(False)

        def _run():
            self._ag_log("info", f"📅 Tổng hợp thư mục {today_str}...")
            # Quét file báo cáo trong thư mục hôm nay (trừ file Tong_hop)
            files = sorted(Path(today_dir).glob('Phieu_xuat_hang_*.xlsx'))
            files = [str(f) for f in files if 'Tong_hop' not in f.name]
            if not files:
                self._ag_log("warn", "⚠ Không tìm thấy file báo cáo nào để tổng hợp")
                self._aggregate_today_btn.setEnabled(True)
                return
            self._ag_log("dim", f"  📄 Tìm thấy {len(files)} file")
            try:
                result = _get_calculator()[2](files, str(today_dir), template)
                self._ag_log("ok", f"  ✓ {result['rows']} SKU | Qty={result['tong_qty']} | "
                                    f"Sold={result['tong_sold']} | Promo={result['tong_promo']}")
                for key, lbl in [('xlsx_report', '📊'), ('pdf_report', '📄')]:
                    fp = result['files'].get(key)
                    if fp and Path(fp).exists():
                        self._add_result(fp, label=lbl)
                        self._ag_log("ok", f"  {lbl} {Path(fp).name}")
                self._ag_log("bold_ok", "✅ Tổng hợp hôm nay hoàn tất")
            except Exception as e:
                self._ag_log("err", f"✗ Lỗi: {e}")
            self._aggregate_today_btn.setEnabled(True)

        threading.Thread(target=_run, daemon=True).start()

    # ═══════════════════════════════════════════════════════
    # BOTTOM BAR
    # ═══════════════════════════════════════════════════════
    def _build_bottom_bar(self):
        bar = QWidget()
        bar.setObjectName("bottomBar")
        layout = QVBoxLayout(bar)
        layout.setContentsMargins(24, 16, 24, 16)
        layout.setSpacing(12)

        self.status_label = QLabel("✅ Hệ thống sẵn sàng")
        self.status_label.setObjectName("statusLabel")
        self.status_label.setAlignment(Qt.AlignCenter)
        self.status_label.setStyleSheet("color: #059669;")

        layout.addWidget(self.status_label)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(12)

        self.sched_btn = QPushButton("▶ CHẠY")
        self.sched_btn.setObjectName("schedBtn")
        self.sched_btn.setCursor(Qt.PointingHandCursor)
        self.sched_btn.clicked.connect(self._on_run_schedule)
        btn_row.addWidget(self.sched_btn, 2)

        self.stop_btn = QPushButton("⏹ DỪNG")
        self.stop_btn.setObjectName("stopBtn")
        self.stop_btn.setCursor(Qt.PointingHandCursor)
        self.stop_btn.setEnabled(False)
        self.stop_btn.clicked.connect(self._on_stop)
        btn_row.addWidget(self.stop_btn, 1)

        layout.addLayout(btn_row)
        return bar

    # ═══════════════════════════════════════════════════════
    # STATUS UPDATES
    # ═══════════════════════════════════════════════════════
    def _update_cookie_status(self):
        self.cookie_row.update_cookie_status()

    def _update_master_status(self):
        self.master_row.update_excel_status()

    def _update_retail_status(self):
        self.retail_row.update_excel_status()

    def _update_template_status(self):
        self.template_row.update_excel_status()

    def _on_cookie_changed(self, path: str):
        self._cookie_real = path
        self._update_cookie_status()

    def _on_master_changed(self, path: str):
        self._master_real = path
        self._update_master_status()
        self._load_preview_data()

    def _on_retail_changed(self, path: str):
        self._retail_real = path
        self._update_retail_status()
        self._load_preview_data()

    def _on_template_changed(self, path: str):
        self._template_real = path
        self._update_template_status()

    def _on_output_dir_changed(self, path: str):
        os.makedirs(path, exist_ok=True)

    # ═══════════════════════════════════════════════════════
    # PREVIEW TABLE LOGIC WITH SEARCH
    # ═══════════════════════════════════════════════════════
    def _load_preview_data(self):
        mp = self._master_real
        if mp and Path(mp).exists():
            self._populate_table(self.master_table, mp, "master")
        rp = self._retail_real
        if rp and Path(rp).exists():
            self._populate_table(self.retail_table, rp, "retail")

    def _populate_table(self, table: QTableWidget, file_path: str, table_type: str):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, data_only=True, read_only=True)
            ws = wb.active

            headers = [str(cell.value) if cell.value else f"Col{i}"
                       for i, cell in enumerate(ws[1], 1)]

            if table_type == "master":
                cols = headers[:6] if len(headers) >= 6 else headers
            else:
                cols = headers[:4] if len(headers) >= 4 else headers

            # Block signals để cellChanged không fire khi đang load dữ liệu
            table.blockSignals(True)

            table.setColumnCount(len(cols))
            table.setHorizontalHeaderLabels(cols)

            rows_data = []
            excel_rows_list = []
            max_rows = 5001  # Nâng hạn mức load lên 5000 dòng
            for i, row in enumerate(ws.iter_rows(min_row=2, max_row=max_rows, values_only=True)):
                if any(cell is not None for cell in row):
                    rows_data.append(row)
                    excel_rows_list.append(i + 2)  # Hàng thực trong Excel (1-indexed)

            table.setRowCount(len(rows_data))
            for ri, row in enumerate(rows_data):
                for ci in range(len(cols)):
                    val = str(row[ci]) if ci < len(row) and row[ci] is not None else ""
                    item = QTableWidgetItem(val)
                    table.setItem(ri, ci, item)

            # Lưu metadata để cellChanged handler biết ghi vào đâu
            table._file_path = file_path
            table._excel_rows = excel_rows_list

            wb.close()
            table.blockSignals(False)
            table.resizeColumnsToContents()
            table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        except Exception:
            pass

    def _on_table_cell_changed(self, row: int, col: int):
        """Khi người dùng sửa một ô → tự động ghi vào file Excel gốc."""
        table = self.sender()
        file_path = getattr(table, '_file_path', None)
        excel_rows = getattr(table, '_excel_rows', [])
        if not file_path or row >= len(excel_rows):
            return
        excel_row = excel_rows[row]  # Hàng thực trong Excel (1-indexed)
        item = table.item(row, col)
        if item is None:
            return
        raw_text = item.text().strip()
        # Tự động đoán kiểu dữ liệu: int → float → string
        if raw_text == "":
            new_value = None
        else:
            try:
                new_value = int(raw_text)
            except ValueError:
                try:
                    new_value = float(raw_text)
                except ValueError:
                    new_value = raw_text
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            ws = wb.active
            ws.cell(row=excel_row, column=col + 1, value=new_value)
            wb.save(file_path)
            wb.close()
        except Exception:
            pass  # File đang mở bởi ứng dụng khác → bỏ qua

    def _on_preview_tab_changed(self, idx: int):
        is_master = (idx == 0)
        self.master_table.setVisible(is_master)
        self.retail_table.setVisible(not is_master)
        
        for i, btn in enumerate([self.preview_master_btn, self.preview_retail_btn]):
            btn.setProperty("active", i == idx)
            btn.style().unpolish(btn)
            btn.style().polish(btn)
            
        for table in [self.master_table, self.retail_table]:
            for row in range(table.rowCount()):
                table.setRowHidden(row, False)
        
        self._filter_preview_table(self.search_input.text())

    def _filter_preview_table(self, text: str):
        text_lower = text.lower().strip()
        active_table = self.master_table if self.preview_btn_group.checkedId() == 0 else self.retail_table
        
        for row in range(active_table.rowCount()):
            if not text_lower:
                active_table.setRowHidden(row, False)
                continue
            match = False
            for col in range(active_table.columnCount()):
                item = active_table.item(row, col)
                if item and text_lower in item.text().lower():
                    match = True
                    break
            active_table.setRowHidden(row, not match)

    # ═══════════════════════════════════════════════════════
    # ADD / DELETE ROWS IN PREVIEW TABLE
    # ═══════════════════════════════════════════════════════
    def _get_active_table(self) -> QTableWidget:
        """Trả về bảng đang hiển thị (master hoặc retail)."""
        return self.master_table if self.preview_btn_group.checkedId() == 0 else self.retail_table

    def _on_add_row(self):
        """Thêm 1 dòng trống vào cuối bảng và ghi vào file Excel gốc."""
        table = self._get_active_table()
        file_path = getattr(table, '_file_path', None)
        excel_rows = getattr(table, '_excel_rows', [])
        if not file_path:
            QMessageBox.warning(self, "Cảnh báo", "Không tìm thấy file Excel để ghi.")
            return

        # Tìm Excel row tiếp theo (sau dòng cuối cùng hiện tại)
        last_excel_row = excel_rows[-1] if excel_rows else 1  # row 1 = header
        new_excel_row = last_excel_row + 1

        # Thêm dòng mới vào bảng
        new_row_idx = table.rowCount()
        table.blockSignals(True)
        table.setRowCount(new_row_idx + 1)
        for ci in range(table.columnCount()):
            table.setItem(new_row_idx, ci, QTableWidgetItem(""))
        table._excel_rows.append(new_excel_row)
        table.blockSignals(False)

        # Ghi dòng trống vào file Excel
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            ws = wb.active
            for ci in range(table.columnCount()):
                ws.cell(row=new_excel_row, column=ci + 1, value=None)
            wb.save(file_path)
            wb.close()
        except Exception:
            pass  # File đang mở → bỏ qua

        # Scroll xuống dòng mới
        table.scrollToBottom()
        table.selectRow(new_row_idx)
        self._log_html("dim", f"  ➕ Đã thêm dòng mới (Excel row {new_excel_row})")

    def _on_delete_rows(self):
        """Xóa các dòng đang chọn khỏi bảng và file Excel gốc."""
        table = self._get_active_table()
        file_path = getattr(table, '_file_path', None)
        excel_rows = getattr(table, '_excel_rows', [])
        if not file_path:
            QMessageBox.warning(self, "Cảnh báo", "Không tìm thấy file Excel để ghi.")
            return

        selected_rows = sorted(set(idx.row() for idx in table.selectedIndexes()), reverse=True)
        if not selected_rows:
            QMessageBox.information(self, "Thông báo", "Chọn ít nhất 1 dòng để xóa.")
            return

        excel_rows_to_delete = [excel_rows[r] for r in selected_rows if r < len(excel_rows)]

        reply = QMessageBox.question(
            self, "Xác nhận xóa",
            f"Xóa {len(selected_rows)} dòng đã chọn?\n\nKhông thể hoàn tác.",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return

        # Xóa trong file Excel trước (theo thứ tự row từ lớn → nhỏ để không lệch)
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            ws = wb.active
            for excel_row in sorted(excel_rows_to_delete, reverse=True):
                ws.delete_rows(excel_row)
            wb.save(file_path)
            wb.close()
        except Exception:
            QMessageBox.warning(self, "Lỗi", "Không thể ghi file Excel. File có thể đang mở.")
            return

        # Xóa trong bảng (từ lớn → nhỏ)
        table.blockSignals(True)
        for row_idx in selected_rows:
            table.removeRow(row_idx)
            del table._excel_rows[row_idx]
        table.blockSignals(False)

        self._log_html("dim", f"  🗑 Đã xóa {len(selected_rows)} dòng")

    # ═══════════════════════════════════════════════════════
    # BUTTON STATE MANAGEMENT
    # ═══════════════════════════════════════════════════════
    def _set_buttons(self, state: str):
        if state in ("running", "scheduled"):
            self.sched_btn.setEnabled(False)
            self.stop_btn.setEnabled(True)
        else:
            self.sched_btn.setEnabled(True)
            self.stop_btn.setEnabled(False)

    # ═══════════════════════════════════════════════════════
    # CONFIG COLLECTOR
    # ═══════════════════════════════════════════════════════
    def _collect_config(self) -> dict:
        carriers_to_process = []
        carrier_keys = ["spx", "ghn", "nv", "vnp", "jt", "best", "vt",
                       "vnp_bulky", "ghn_bulky", "njv_bulky", "spx_bulky", "vtp_bulky",
                       "ahamove", "ahamove_sbs", "spx_tn", "spx_sbs", "spx_locker", "other"]
        carrier_names = ["SPX Express", "Giao Hàng Nhanh", "Ninja Van", "VNPost Nhanh",
                        "J&T Express", "BEST Express", "Viettel Post",
                        "VNP - Hàng Cồng Kềnh", "GHN - Hàng Cồng Kềnh", "NJV - Hàng Cồng Kềnh",
                        "SPX - Hàng Cồng Kềnh", "VTP - Hàng Cồng Kềnh",
                        "Ahamove - Trong Ngày", "Ahamove SBS - Trong Ngày",
                        "SPX Express - Trong Ngày", "SPX Express SBS - Trong Ngày",
                        "Tủ nhận hàng - SPX", "Đơn vị vận chuyển khác"]
        for key, name in zip(carrier_keys, carrier_names):
            if self.carrier_checkboxes[key].isChecked():
                val = self.carrier_spinboxes[key].value()
                carriers_to_process.append((name, val))
            # unchecked = bỏ qua hãng này hoàn toàn

        return {
            "cookie": self._cookie_real,
            "output_dir": self.output_row.get_real_path() or str(BASE_DIR / "outputs"),
            "master": self._master_real,
            "retail": self._retail_real,
            "template": self._template_real,
            "carriers": carriers_to_process,
            "auto_print": self.auto_print_cb.isChecked(),
            "printer": self.printer_combo.currentText(),
            "test_mode": self.test_mode_cb.isChecked(),
            "exclude_pre_orders": self.exclude_pre_orders_cb.isChecked(),
            "batch_size": self.batch_size_spin.value(),
            "pdf_settings": self._build_pdf_settings(),
        }

    def _build_pdf_settings(self) -> str:
        parts = ["paper=A4",
                 f"duplex={self.duplex_combo.currentText()}"]
        return ",".join(parts)

    # ═══════════════════════════════════════════════════════
    # PRINT HELPERS
    # ═══════════════════════════════════════════════════════
    def _refresh_printers(self):
        printers = _get_printers()
        self.printer_combo.clear()
        self.printer_combo.addItems(printers)
        if printers:
            default_printer = _get_default_printer()
            if default_printer and default_printer in printers:
                self.printer_combo.setCurrentIndex(printers.index(default_printer))
            else:
                self.printer_combo.setCurrentIndex(0)

    # ═══════════════════════════════════════════════════════
    # LOG (colored HTML via QTextEdit)
    # ═══════════════════════════════════════════════════════
    def _log_html(self, tag: str, msg: str):
        ts = datetime.now().strftime("%H:%M:%S")
        color = self.TAG_COLORS.get(tag, "#E2E8F0")
        weight = "font-weight: bold;" if "bold" in tag else ""
        html = (
            f"<span style='color:{color};'>{ts}  </span>"
            f"<span style='color:{color};{weight}'>{msg}</span><br>"
        )
        cursor = self.log_view.textCursor()
        cursor.movePosition(QTextCursor.End)
        cursor.insertHtml(html)
        # Giới hạn ~5000 dòng: xóa nửa đầu thay vì clear toàn bộ
        doc = self.log_view.document()
        if doc.blockCount() > 5000:
            remove_count = doc.blockCount() - 2500
            cursor = QTextCursor(doc.begin())
            for _ in range(remove_count):
                cursor.movePosition(QTextCursor.NextBlock, QTextCursor.KeepAnchor)
            cursor.removeSelectedText()
            # Thông báo đã trim log
            cursor.movePosition(QTextCursor.Start)
            cursor.insertHtml(
                f"<span style='color:#64748B;'>[... đã xóa {remove_count} dòng cũ — giữ 2500 dòng gần nhất ...]</span><br>"
            )
            self.log_view.ensureCursorVisible()

    @Slot(str, str)
    def _on_log_message(self, tag: str, msg: str):
        self._log_html(tag, msg)

    @Slot(str, str)
    def _on_state_changed(self, step: str, msg: str):
        colors = {
            "idle": "#059669", "running": "#D97706",
            "waiting": "#2563EB", "done": "#059669", "error": "#DC2626",
        }
        color = colors.get(step, "#1E293B")
        self.status_label.setText(msg)
        self.status_label.setStyleSheet(f"color: {color}; font-weight: bold; font-size: 14px;")

    @Slot(object)
    def _on_job_completed(self, result: dict):
        self.running = False
        # Ghi nhận thời điểm hoàn thành cho scheduler interval (tính từ lúc KẾT THÚC)
        if self.scheduler_active:
            self._sched_last_run = datetime.now()
        else:
            self._set_buttons("idle")
        if result:
            # PDF paths đã được thêm qua result_file signal trong worker — không thêm lại
            for r in result.get('results', []):
                for key, lbl in [('xlsx_report', '📊'), ('pdf_report', '📄')]:
                    fp = r['files'].get(key) if isinstance(r.get('files'), dict) else None
                    if fp and Path(fp).exists():
                        self._add_result(fp, label=lbl)

    def _clear_log(self):
        self.log_view.clear()

    # ═══════════════════════════════════════════════════════
    # RESULTS
    # ═══════════════════════════════════════════════════════
    def _add_result(self, file_path: str, label: str = ''):
        """Thêm file vào danh sách kết quả. file_path là đường dẫn đầy đủ."""
        display = f'{label} {Path(file_path).name}' if label else f'📄 {Path(file_path).name}'
        self.result_list.addItem(display)
        self.result_list.item(self.result_list.count() - 1).setData(Qt.UserRole, file_path)
        self.result_files.append(file_path)

    def _clear_results(self):
        self.result_list.clear()
        self.result_files.clear()

    def _on_open_result(self, item: QListWidgetItem):
        fp = item.data(Qt.UserRole)
        if fp and Path(fp).exists():
            os.startfile(fp)

    def _open_output_dir(self):
        base = self.output_row.get_real_path() or str(BASE_DIR / "outputs")
        d = Path(base) / datetime.now().strftime("%Y-%m-%d")
        if d.exists():
            os.startfile(str(d))
        elif Path(base).exists():
            os.startfile(base)

    # ═══════════════════════════════════════════════════════
    # ACTION CALLBACKS
    # ═══════════════════════════════════════════════════════
    def _on_run_now(self):
        if self.running:
            return
        if not Path(self._cookie_real).exists():
            QMessageBox.critical(self, "Lỗi", "Chọn file cookie JSON hợp lệ.")
            return
        config = self._collect_config()
        if not config['carriers']:
            QMessageBox.warning(self, "Cảnh báo", "Vui lòng chọn ít nhất 1 hãng vận chuyển để in.")
            return

        # ── Tính output_dir với date subfolder ──
        base_dir = config['output_dir']
        today_str = datetime.now().strftime('%Y-%m-%d')
        out_dir = str(Path(base_dir) / today_str)
        os.makedirs(out_dir, exist_ok=True)
        config['output_dir'] = out_dir  # Ghi đè = path có ngày

        self.running = True
        self._set_buttons("running")
        self._clear_log()
        self._log_html("bold_ok", "▶ Bắt đầu...")
        self.status_label.setText("⏳ Đang xử lý...")
        self.status_label.setStyleSheet("color: #D97706; font-weight: bold; font-size: 14px;")
        self.trigger_job.emit(config)

    def _on_run_schedule(self):
        if not Path(self._cookie_real).exists():
            QMessageBox.critical(self, "Lỗi", "Chọn file cookie JSON hợp lệ.")
            return
        mode = self._sched_mode
        if mode == "once":
            self._on_run_now()
            return
        if mode == "weekly":
            # Parse per-day config
            day_names = ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7", "Chủ Nhật"]
            self._sched_weekly_config = {}
            has_any = False

            for idx in range(7):
                if not self.weekly_day_checkboxes[idx].isChecked():
                    continue

                time_text = self.weekly_day_time_edits[idx].text().strip()
                if not time_text:
                    continue  # checked but no times entered → skip silently

                parts = [t.strip() for t in time_text.split(",") if t.strip()]
                parsed = []
                for t in parts:
                    try:
                        h, m = t.split(":")
                        h_int, m_int = int(h), int(m)
                        if not (0 <= h_int <= 23 and 0 <= m_int <= 59):
                            raise ValueError
                        parsed.append((h_int, m_int))
                    except (ValueError, TypeError):
                        QMessageBox.critical(self, "Lỗi",
                            f"Giờ không hợp lệ cho {day_names[idx]}: '{t}'. Nhập dạng HH:MM (0-23:0-59).")
                        return

                if parsed:
                    self._sched_weekly_config[idx] = sorted(parsed)
                    has_any = True

            if not has_any:
                QMessageBox.critical(self, "Lỗi",
                    "Vui lòng chọn ít nhất một ngày và nhập ít nhất một khung giờ.")
                return
        elif mode == "interval":
            self._sched_interval_hours = self.interval_spin.value()

        self._clear_log()
        self._clear_results()
        self._log_html("info", f"⏰ Hẹn giờ: {mode}")
        if mode == "interval":
            self._log_html("info", f"   Chạy mỗi {self._sched_interval_hours} giờ")
        elif mode == "weekly":
            day_names = ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6", "Thứ 7", "Chủ Nhật"]
            for idx in range(7):
                if idx in self._sched_weekly_config:
                    times_str = ", ".join(f"{h:02d}:{m:02d}" for h, m in self._sched_weekly_config[idx])
                    self._log_html("info", f"   {day_names[idx]}: {times_str}")

        self._sched_mode = mode
        self.scheduler_active = True
        self._sched_last_run = None
        self._set_buttons("scheduled")
        self._sched_timer.start()
        self._update_countdown()

    def _on_stop(self):
        self.scheduler_active = False
        self._sched_timer.stop()
        self.running = False
        # Dùng invokeMethod để gửi lệnh stop qua event queue của worker thread (đúng chuẩn Qt)
        # threading.Event.set() là thread-safe nên nếu invokeMethod thất bại, direct call vẫn an toàn
        if not QMetaObject.invokeMethod(self._worker, "stop_job", Qt.QueuedConnection):
            self._worker.stop_job()  # fallback an toàn vì chỉ set threading.Event
        self._set_buttons("idle")
        self.status_label.setText("⏹ Đã dừng")
        self.status_label.setStyleSheet("color: #DC2626; font-weight: bold; font-size: 14px;")
        self._log_html("warn", "⏹ Đã dừng hệ thống")

    def _on_schedule_mode_changed(self, btn: QRadioButton):
        mode = btn.property("mode")
        self._sched_mode = mode
        self.interval_panel.setVisible(mode == "interval")
        self.weekly_panel.setVisible(mode == "weekly")

    # ═══════════════════════════════════════════════════════
    # WEEKLY SCHEDULER HELPERS
    # ═══════════════════════════════════════════════════════
    def _on_weekly_apply_all(self):
        """Copy nội dung ô giờ mẫu vào tất cả các ngày đang checked."""
        master_text = self.weekly_master_time_edit.text()
        for idx in range(7):
            if self.weekly_day_checkboxes[idx].isChecked():
                self.weekly_day_time_edits[idx].setText(master_text)

    def _find_next_run_today(self, now: datetime) -> datetime | None:
        """Tìm timeslot tiếp theo trong ngày hôm nay. Trả về None nếu hôm nay hết slot."""
        day_idx = now.weekday()  # Python: 0=Monday → khớp với idx của ta
        slots = self._sched_weekly_config.get(day_idx, [])
        best = None
        for h, m in slots:
            candidate = now.replace(hour=h, minute=m, second=0, microsecond=0)
            if candidate > now and (best is None or candidate < best):
                best = candidate
        return best

    def _calc_next_weekly_run(self, from_time: datetime) -> datetime | None:
        """Quét tối đa 8 ngày tới, tìm timeslot sớm nhất. Dùng cho countdown display."""
        best = None
        for offset in range(8):
            check_date = from_time.date() + timedelta(days=offset)
            day_idx = check_date.weekday()
            day_slots = self._sched_weekly_config.get(day_idx, [])
            for h, m in day_slots:
                candidate = datetime(check_date.year, check_date.month, check_date.day, h, m, 0, 0)
                if candidate > from_time and (best is None or candidate < best):
                    best = candidate
            # Nếu đã tìm thấy slot trong ngày đang xét thì dừng (không cần quét tiếp)
            if best is not None and best.date() == check_date:
                break
        return best

    # ═══════════════════════════════════════════════════════
    # SCHEDULER (QTimer-based, main thread)
    # ═══════════════════════════════════════════════════════
    def _check_schedule(self):
        if not self.scheduler_active:
            return
        now = datetime.now()
        if self._sched_mode == "interval":
            # Lần đầu chạy ngay, các lần sau cách nhau N giờ tính từ lúc hoàn thành
            if self._sched_last_run is None:
                self._sched_last_run = now
                self._execute_scheduled_job()
                return
            next_run = self._sched_last_run + timedelta(hours=self._sched_interval_hours)
            self._sched_next_run = next_run
            if now >= next_run:
                self._execute_scheduled_job()
            else:
                self._update_countdown()
        elif self._sched_mode == "weekly":
            now = datetime.now()
            next_today = self._find_next_run_today(now)
            if next_today is not None:
                self._sched_next_run = next_today
                diff = (next_today - now).total_seconds()
                if diff <= 1:
                    if not self._sched_last_run or (now - self._sched_last_run).total_seconds() > 60:
                        self._execute_scheduled_job()
                else:
                    self._update_countdown()
            else:
                self._sched_next_run = self._calc_next_weekly_run(now)
                self._update_countdown()

    def _execute_scheduled_job(self):
        if self.running:
            self._log_html("dim", "⏭ Bỏ qua chu kỳ — job trước vẫn đang chạy")
            return
        config = self._collect_config()
        if not config['carriers']:
            self._log_html("warn", "⏭ Bỏ qua chu kỳ — không có hãng nào được chọn")
            return

        # ── Tính output_dir với date subfolder ──
        base_dir = config['output_dir']
        today_str = datetime.now().strftime('%Y-%m-%d')
        out_dir = str(Path(base_dir) / today_str)
        os.makedirs(out_dir, exist_ok=True)
        config['output_dir'] = out_dir

        self.running = True
        self._clear_results()
        self.status_label.setText("🔄 Đang chạy tác vụ tự động...")
        self.status_label.setStyleSheet("color: #D97706; font-weight: bold; font-size: 14px;")
        self.trigger_job.emit(config)

    def _update_countdown(self):
        if self._sched_next_run:
            remaining = self._sched_next_run - datetime.now()
            secs = max(0, int(remaining.total_seconds()))
            h, m = secs // 3600, (secs % 3600) // 60
            self.status_label.setText(f"⏳ Chạy tiếp sau {h}h{m:02d}")
            self.status_label.setStyleSheet("color: #2563EB; font-weight: bold; font-size: 14px;")

    # ═══════════════════════════════════════════════════════
    # CLOSE EVENT
    # ═══════════════════════════════════════════════════════
    def closeEvent(self, event):
        self.scheduler_active = False
        self._sched_timer.stop()
        self.running = False
        # shutdown() chứa browser.close() + playwright.stop() là process-level, không phụ thuộc thread
        # Gọi trực tiếp là an toàn; invokeMethod + BlockingQueuedConnection có thể deadlock nếu worker đang bận
        self._worker.shutdown()
        self._worker_thread.quit()
        self._worker_thread.wait(5000)
        event.accept()


# ============================================================
# ENTRY POINT
# ============================================================
if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    window = App()
    window.show()
    sys.exit(app.exec())