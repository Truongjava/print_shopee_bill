"""
calculator.py — Module xử lý PDF danh sách sản phẩm
==================================================
Đối chiếu Picking List PDF với master_data.xlsx để tính số lượng thực tế.
Mỗi Seller SKU trong PDF được tra cứu trong master_data, sau đó nhân số lượng.

Hỗ trợ 2 định dạng PDF:
  - TikTok Picking List (regex pattern: SKU + Qty + OrderID 15+ digits)
  - Shopee Phiếu xuất hàng  (column-position parsing: cột SKU phân loại + Số lượng + Order SN)
"""
import os
import re
from collections import defaultdict
from datetime import datetime

from fpdf import FPDF
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# MASTER DATA
# ============================================================

def load_master_data(master_path: str) -> list[dict]:
    """
    Đọc file master_data.xlsx (hoặc mã combo.xlsx).
    Tự động nhận diện cột theo tên header, không phụ thuộc vị trí.
    Trả về list các dòng, mỗi dòng là dict:
      {seller_sku, sku, qty, qty_sold, promo_qty}
    """
    from openpyxl import load_workbook

    wb = load_workbook(master_path, data_only=True)
    ws = wb.active

    # ── Đọc header để xác định vị trí các cột ──
    headers = {}
    for ci, cell in enumerate(ws[1], 1):
        val = str(cell.value).lower().strip() if cell.value else ''
        headers[ci] = val

    def find_col(keywords: list[str]) -> int:
        """Tìm cột theo từ khóa (không phân biệt hoa thường)."""
        for ci, h in headers.items():
            for kw in keywords:
                if kw in h:
                    return ci
        return 0

    col_seller_sku = find_col(['combo', 'seller sku', 'seller_sku', 'mã combo', 'ma combo']) or 1
    col_sku = find_col(['sku']) or 2
    col_qty = find_col(['sl', 'qty', 'số lượng', 'so luong']) or 3
    col_qty_sold = find_col(['sl bán', 'sl ban', 'qty sold', 'qty_sold', 'bán', 'ban']) or 4
    col_promo = find_col(['sl km', 'promo qty', 'promo_qty', 'km']) or 5
    col_unit = find_col(['đơn vị tính', 'don vi tinh', 'đơn vị', 'don vi', 'dvt', 'unit']) or 0

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[col_seller_sku - 1] is None:
            continue
        seller_sku = str(row[col_seller_sku - 1]).strip()
        sku = str(row[col_sku - 1]).strip() if row[col_sku - 1] is not None else seller_sku
        qty = int(row[col_qty - 1]) if row[col_qty - 1] is not None else 0
        qty_sold = int(row[col_qty_sold - 1]) if row[col_qty_sold - 1] is not None else 0
        promo_qty = int(row[col_promo - 1]) if row[col_promo - 1] is not None else 0
        unit = str(row[col_unit - 1]).strip() if col_unit and row[col_unit - 1] is not None else ''
        rows.append({
            "seller_sku": seller_sku,
            "sku": sku,
            "qty": qty,
            "qty_sold": qty_sold,
            "promo_qty": promo_qty,
            "unit": unit,
        })
    wb.close()
    return rows


def load_retail_data(retail_path: str) -> dict[str, dict]:
    """
    Doc file san pham ban le.xlsx.
    Cot: SKU | Don vi tinh | SL | SL ban | SL KM
    Khong co cot COMBO -> Seller SKU chinh la SKU.
    Tra ve dict: {sku: {seller_sku, sku, unit, qty, qty_sold, promo_qty}}
    """
    from openpyxl import load_workbook

    wb = load_workbook(retail_path, data_only=True)
    ws = wb.active

    headers = {}
    for ci, cell in enumerate(ws[1], 1):
        val = str(cell.value).lower().strip() if cell.value else ''
        headers[ci] = val

    def find_col(keywords):
        for ci, h in headers.items():
            for kw in keywords:
                if kw in h:
                    return ci
        return 0

    col_sku = find_col(['sku', 'mã sản phẩm', 'ma san pham', 'mã sản phẩm']) or 1
    col_name = find_col(['tên sản phẩm', 'ten san pham', 'tên sp', 'product name']) or 0
    col_unit = find_col(['đơn vị tính', 'don vi tinh', 'đơn vị', 'don vi', 'dvt', 'unit']) or 0
    col_qty = find_col(['sl', 'qty', 'số lượng', 'so luong']) or 3
    col_sold = find_col(['sl bán', 'sl ban', 'qty sold', 'bán', 'ban']) or 5
    col_promo = find_col(['sl km', 'promo qty', 'khuyến mại', 'khuyen mai', 'km']) or 6

    retail = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = str(row[col_sku - 1]).strip() if row[col_sku - 1] is not None else ''
        if not sku:
            continue
        retail[sku] = {
            "seller_sku": sku,
            "sku": sku,
            "product_name": str(row[col_name - 1]).strip() if col_name and row[col_name - 1] is not None else '',
            "unit": str(row[col_unit - 1]).strip() if col_unit and row[col_unit - 1] is not None else '',
            "qty": int(row[col_qty - 1]) if row[col_qty - 1] is not None else 1,
            "qty_sold": int(row[col_sold - 1]) if row[col_sold - 1] is not None else 1,
            "promo_qty": int(row[col_promo - 1]) if row[col_promo - 1] is not None else 0,
        }
    wb.close()
    print(f"   Retail data: {len(retail)} SKUs")
    return retail


def build_sku_index(master_data: list[dict]) -> tuple[set[str], dict[str, str], dict[str, list[str]]]:
    """
    Từ master_data trả về:
      - master_skus: set tất cả Seller SKU
      - prefix_map: map từ SKU bị ngắt dòng → full Seller SKU (chỉ khi unique)
        VD: "CER01-" → "CER01-BOG17-2"
      - ambiguous_map: map từ prefix ambiguous → list các full SKU có thể
        VD: "ANTQ-3-" → ["ANTQ-3-6MET01", "ANTQ-3-TRG01"]
    """
    master_skus = set(r["seller_sku"] for r in master_data)

    prefix_map: dict[str, str] = {}
    ambiguous_map: dict[str, list[str]] = {}
    for sku in master_skus:
        if '-' in sku:
            parts = sku.split('-')
            for i in range(1, len(parts)):
                prefix = '-'.join(parts[:i]) + '-'
                if prefix not in prefix_map and prefix not in ambiguous_map:
                    prefix_map[prefix] = sku
                elif prefix in prefix_map:
                    # Trở thành ambiguous
                    ambiguous_map[prefix] = [prefix_map[prefix], sku]
                    del prefix_map[prefix]
                else:
                    ambiguous_map[prefix].append(sku)

    return master_skus, prefix_map, ambiguous_map


# ============================================================
# PDF TYPE DETECTION
# ============================================================

def detect_pdf_type(pdf_path: str) -> str:
    """
    Tự động nhận diện loại PDF: 'shopee' | 'tiktok' | 'unknown'
    Dựa vào text header của trang đầu tiên.
    """
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if not pdf.pages:
                return 'unknown'
            first_page_text = pdf.pages[0].extract_text() or ''
    except Exception:
        return 'unknown'

    # Shopee: "Phiếu xuất hàng" + "Order SN" (cột Order SN thay vì Order ID)
    if 'phiếu xuất hàng' in first_page_text.lower():
        return 'shopee'
    if 'order sn' in first_page_text.lower():
        return 'shopee'

    # TikTok: "Picking list" hoặc "Order quantity:" + Order ID dạng số dài
    if 'picking list' in first_page_text.lower():
        return 'tiktok'
    if re.search(r'Order quantity:\s*\d+', first_page_text):
        return 'tiktok'

    # Fallback: thử tìm pattern TikTok (Order ID 15+ chữ số)
    if re.search(r'\b\d{15,}\b', first_page_text):
        return 'tiktok'

    # Thử tìm pattern Shopee (Order SN 12 ký tự: 6 số + 6 chữ-số)
    if re.search(r'\b\d{6}[A-Z0-9]{4,}\b', first_page_text):
        return 'shopee'

    return 'unknown'


def split_shopee_pdf(pdf_path: str, output_dir: str = '') -> tuple[str, str, int]:
    """
    Tách 1 file Shopee PDF gộp (phiếu xuất + shipping label) thành 2 file riêng:
      - Phiếu xuất hàng (picking list): các trang có header "# SKU" hoặc "Phiếu xuất hàng"
      - Shipping label: các trang có "Mã vận đơn" hoặc "THÔNG TIN ĐƠN HÀNG"

    Trả về: (picking_path, shipping_path)
    Nếu không tìm thấy loại nào thì trả về '' cho loại đó.
    """
    from pypdf import PdfReader, PdfWriter

    # ── Guard: nếu file đã được tách rồi thì không tách lại nữa ──
    fname_lower = os.path.basename(pdf_path).lower()
    if '_phieu_xuat' in fname_lower or '_shipping_label' in fname_lower:
        print(f'   ⏭ Bỏ qua tách: {os.path.basename(pdf_path)} (đã được tách trước đó)')
        return '', '', 0

    # ── Dùng pdfplumber để phân loại trang (xử lý tiếng Việt đúng) ──
    picking_indices: list[int] = []
    shipping_indices: list[int] = []

    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages):
            # Chỉ cần đọc text của trang đầu để phân loại (đủ nhanh)
            text = page.extract_text() or ''

            is_picking = (
                'phiếu xuất hàng' in text.lower() or
                ('# SKU' in text and 'SKU phân loại' in text)
            )
            is_shipping = (
                'mã vận đơn' in text.lower() or
                'THÔNG TIN ĐƠN HÀNG' in text
            )

            if is_picking:
                picking_indices.append(i)
            elif is_shipping:
                shipping_indices.append(i)

    # ── Dùng pypdf để tách file (nhanh, không cần đọc text) ──
    reader = PdfReader(pdf_path)
    picking_writer = PdfWriter()
    shipping_writer = PdfWriter()

    for idx in picking_indices:
        picking_writer.add_page(reader.pages[idx])
    for idx in shipping_indices:
        shipping_writer.add_page(reader.pages[idx])
        # Bỏ qua các trang không xác định được (nếu có)

    # ── Lưu file ──
    out_dir = output_dir or os.path.dirname(pdf_path)
    base_name = os.path.splitext(os.path.basename(pdf_path))[0]

    picking_path = ''
    shipping_path = ''

    if picking_indices:
        picking_path = os.path.join(out_dir, f'{base_name}_phieu_xuat.pdf')
        picking_writer.write(picking_path)
        print(f'   📋 Đã tách {len(picking_indices)} trang phiếu xuất → {os.path.basename(picking_path)}')

    if shipping_indices:
        shipping_path = os.path.join(out_dir, f'{base_name}_shipping_label.pdf')
        shipping_writer.write(shipping_path)
        print(f'   📦 Đã tách {len(shipping_indices)} trang shipping label → {os.path.basename(shipping_path)}')

    return picking_path, shipping_path, len(shipping_indices)


# ============================================================
# PDF EXTRACTION — SHOPEE (Phiếu xuất hàng)
# ============================================================

def _parse_shopee_header(full_text: str) -> dict:
    """
    Trích xuất thông tin header từ Shopee Phiếu xuất hàng PDF.
    VD: Ngày in phiếu:03:37 PM 12/06/2026
    Trả về: {print_time, total_orders}
    """
    info = {}
    m = re.search(r'Ngày in phiếu:\s*(.+)', full_text)
    if m:
        info['print_time'] = m.group(1).strip()

    # Đếm tổng số Order SN duy nhất từ picking list
    order_sns = set(re.findall(r'\b(\d{6}[A-Z0-9]{4,10})\b', full_text))
    info['order_qty'] = len(order_sns)
    info['product_qty'] = len(order_sns)
    info['item_qty'] = len(order_sns)

    return info


def _extract_order_sns_from_shipping(shipping_pdf_path: str) -> list[str]:
    """
    Trích xuất Order SN ĐẦY ĐỦ từ các trang shipping label.
    Shipping label có dòng: Mã đơn hàng: 2607313WXH0RK0
    Trả về list các Order SN đã sắp xếp.
    """
    import pdfplumber
    full_sns: set[str] = set()
    try:
        with pdfplumber.open(shipping_pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ''
                # Pattern: Mã đơn hàng: XXXX (Order SN đầy đủ, thường 12-16 ký tự)
                found = re.findall(r'Mã đơn hàng:\s*(\d{6}[A-Z0-9]{4,10})', text)
                full_sns.update(found)
    except Exception:
        pass
    return sorted(full_sns)


def _group_words_by_row(pdf_path: str, picking_only: bool = True) -> list[list[dict]]:
    """
    Trích xuất tất cả words từ PDF, nhóm theo dòng (Y position ~20px).
    Nếu picking_only=True, chỉ lấy từ trang phiếu xuất (bỏ qua shipping label).
    Trả về list các dòng, mỗi dòng là list words đã sắp xếp theo X.
    """
    from collections import defaultdict

    def _is_picking_page(text: str) -> bool:
        return (
            'phiếu xuất hàng' in text.lower() or
            ('# SKU' in text and 'SKU phân loại' in text)
        )

    def _is_shipping_page(text: str) -> bool:
        return (
            'mã vận đơn' in text.lower() or
            'THÔNG TIN ĐƠN HÀNG' in text
        )

    all_lines = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            words = page.extract_words(keep_blank_chars=True, x_tolerance=3)
            if not words:
                continue

            # Kiểm tra loại trang
            if picking_only:
                page_text = page.extract_text() or ''
                if _is_shipping_page(page_text) and not _is_picking_page(page_text):
                    continue  # Bỏ qua trang shipping label

            # Nhóm từ theo Y (làm tròn 10px — tránh banker's rounding gộp dòng)
            rows = defaultdict(list)
            for w in words:
                row_y = round(w['top'] / 10) * 10
                rows[row_y].append(w)

            for y in sorted(rows.keys()):
                line_words = sorted(rows[y], key=lambda w: w['x0'])
                all_lines.append({
                    'page': page.page_number,
                    'y': y,
                    'words': line_words,
                })

    return all_lines


def _extract_shopee_lines(pdf_path: str) -> list[dict]:
    """
    Dùng extract_chars() thay vì extract_words() để tránh lỗi merge text
    giữa các cột. Chỉ quét cột SKU phân loại (x bắt đầu ~326) và Số lượng (x~463).

    Trả về list các dòng: {page, y, sku, qty}
    """
    from collections import defaultdict

    def _is_shipping_page(text: str) -> bool:
        return (
            'mã vận đơn' in text.lower() or
            'THÔNG TIN ĐƠN HÀNG' in text
        )

    # ── Hằng số vị trí cột (dựa trên vị trí header "SKU phân loại" ở x=326) ──
    SKU_COL_START = 324   # Bắt đầu cột SKU phân loại
    SKU_COL_END = 390     # Kết thúc cột SKU phân loại (trước "Phân loại hàng" ở x=394)
    QTY_COL_START = 460   # Bắt đầu cột Số lượng
    QTY_COL_END = 475     # Kết thúc cột Số lượng

    # ── Từ khóa cần bỏ qua ──
    SKIP_TOKENS = {
        'sku phân loại', 'sku phan loai', 'phân loại hàng', 'phan loai hang',
        'phân loại', 'phan loai', 'số lượng', 'so luong',
    }

    all_lines: list[dict] = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ''

            # Bỏ qua trang shipping label
            if _is_shipping_page(page_text):
                continue

            chars = page.chars
            if not chars:
                continue

            # Gom ký tự theo dòng Y (làm tròn 5px — đủ mịn để không gộp 2 dòng sát nhau)
            rows: dict[int, list] = defaultdict(list)
            for c in chars:
                row_y = round(c['top'] / 5) * 5
                rows[row_y].append(c)

            for y in sorted(rows.keys()):
                line_chars = sorted(rows[y], key=lambda c: c['x0'])

                # ── Trích xuất SKU từ cột SKU phân loại ──
                sku_chars = []
                in_sku_zone = False
                for c in line_chars:
                    if SKU_COL_START <= c['x0'] <= SKU_COL_START + 4:
                        in_sku_zone = True
                    if in_sku_zone and c['x0'] >= SKU_COL_END:
                        break
                    if in_sku_zone:
                        sku_chars.append(c)

                sku_text = ''.join(c['text'] for c in sku_chars).strip()

                # ── Trích xuất Số lượng ──
                qty_chars = []
                in_qty_zone = False
                for c in line_chars:
                    if QTY_COL_START <= c['x0'] <= QTY_COL_START + 6:
                        in_qty_zone = True
                    if in_qty_zone and c['x0'] >= QTY_COL_END:
                        break
                    if in_qty_zone:
                        qty_chars.append(c)

                qty_text = ''.join(c['text'] for c in qty_chars).strip()

                # ── Làm sạch SKU ──
                if sku_text:
                    sku_lower = sku_text.lower()
                    # Bỏ qua dòng header
                    if any(skip in sku_lower for skip in SKIP_TOKENS):
                        sku_text = ''
                    # Chỉ giữ ký tự hợp lệ trong SKU: A-Z, a-z, 0-9, -
                    sku_text = re.sub(r'[^A-Za-z0-9-]', '', sku_text)

                # ── Làm sạch Qty: chỉ giữ chữ số ──
                if qty_text:
                    qty_text = re.sub(r'[^0-9]', '', qty_text)

                # Chỉ thêm dòng nếu có ít nhất SKU hoặc Qty
                if sku_text or qty_text:
                    all_lines.append({
                        'page': page.page_number,
                        'y': y,
                        'sku': sku_text,
                        'qty': qty_text,
                    })

    return all_lines


def extract_order_counts_shopee(
    pdf_path: str,
    master_skus: set[str],
    prefix_map: dict[str, str],
    ambiguous_map: dict[str, list[str]],
    retail_lookup: dict[str, dict] | None = None,
) -> tuple[dict[str, int], dict]:
    """
    Trích xuất số đơn hàng cho mỗi Seller SKU từ Shopee Phiếu xuất hàng PDF.

    Cách tiếp cận: dùng extract_chars() (thay vì extract_words) để tránh lỗi
    merge text giữa các cột. Chỉ quét cột SKU phân loại (x=326, thẳng hàng với
    header) và cột Số lượng (x=463). Bỏ qua cột SKU chính (x=29) vì dễ gây
    false positive khi SKU dài bị ngắt dòng.

    Cột trong Shopee PDF:
      - # (x ~ 15)
      - SKU chính (x ~ 29) — KHÔNG DÙNG, dễ false positive
      - Tên sản phẩm (x ~ 119)
      - SKU phân loại (x ~ 326-390) ← CHỈ DÙNG CỘT NÀY
      - Phân loại hàng (x ~ 394-458)
      - Số lượng (x ~ 463)
      - Order SN (x ~ 511+)

    Trả về: ({seller_sku: tong_so_qty}, header_info)
    """
    lines = _extract_shopee_lines(pdf_path)

    # Tập hợp text đầy đủ để parse header (từ tất cả trang)
    full_text = ' '.join(
        line['sku'] + ' ' + line['qty']
        for line in lines
    )
    header_info = _parse_shopee_header(full_text)

    counts: dict[str, int] = defaultdict(int)

    # ── Xử lý tuần tự từng dòng ──
    pending_prefix = None  # SKU kết thúc bằng "-" cần ghép với dòng sau
    pending_qty = None     # Qty từ dòng có pending_prefix

    for line in lines:
        sku = line['sku']
        qty_str = line['qty']

        # ── Trường hợp 1: Dòng này là continuation của pending prefix ──
        # VD: pending_prefix="ANT01-3-" + sku="6MET01" → "ANT01-3-6MET01"
        if sku and pending_prefix:
            combined = pending_prefix + sku
            if (combined in master_skus or
                combined in prefix_map or
                (retail_lookup and combined in retail_lookup)):
                # Ghép thành công → flush với Qty
                final_qty = qty_str if qty_str else (pending_qty or '1')
                _match_and_add(counts, combined, int(final_qty),
                               master_skus, prefix_map, ambiguous_map, retail_lookup)
                pending_prefix = None
                pending_qty = None
                continue
            elif not qty_str:
                # Combined chưa khớp + không có Qty → tiếp tục accumulate
                # VD: "ANT01-3-" + "6MET" + "01" → cần thêm "01" nữa
                pending_prefix = combined
                continue
            else:
                # Combined không khớp NHƯNG dòng này có Qty riêng
                # → flush pending như 1 SKU độc lập, rồi xử lý tiếp sku hiện tại
                _match_and_add(counts, pending_prefix, int(pending_qty or '1'),
                               master_skus, prefix_map, ambiguous_map, retail_lookup)
                pending_prefix = None
                pending_qty = None
                # Không continue — xử lý sku hiện tại bên dưới

        # ── Trường hợp 2: Dòng có Qty nhưng không có SKU (VD: Qty ở dòng tiếp theo của continuation) ──
        if not sku and qty_str and pending_prefix:
            # Qty này thuộc về pending_prefix
            _match_and_add(counts, pending_prefix, int(qty_str),
                           master_skus, prefix_map, ambiguous_map, retail_lookup)
            pending_prefix = None
            pending_qty = None
            continue

        # ── Trường hợp 3: Dòng có SKU đầy đủ + Qty ──
        if sku:
            if sku.endswith('-'):
                # SKU dạng prefix → lưu lại chờ dòng continuation
                pending_prefix = sku
                pending_qty = qty_str if qty_str else None
            elif qty_str:
                # Dòng đầy đủ: SKU + Qty
                _match_and_add(counts, sku, int(qty_str),
                               master_skus, prefix_map, ambiguous_map, retail_lookup)
                pending_prefix = None
                pending_qty = None
            elif pending_prefix:
                # SKU không kết thúc bằng "-" nhưng có pending_prefix
                # → continuation của pending (VD: "01" sau "6MET")
                combined = pending_prefix + sku
                final_qty = pending_qty or '1'
                _match_and_add(counts, combined, int(final_qty),
                               master_skus, prefix_map, ambiguous_map, retail_lookup)
                pending_prefix = None
                pending_qty = None
            # else: SKU không có Qty và không có pending → bỏ qua
            # (đây chính là các dòng rác như BOG6 ở cột SKU chính)

    # ── Flush pending cuối cùng (nếu còn) ──
    if pending_prefix:
        _match_and_add(counts, pending_prefix, int(pending_qty or '1'),
                       master_skus, prefix_map, ambiguous_map, retail_lookup)

    return dict(counts), header_info


def _match_and_add(
    counts: dict[str, int],
    candidate: str,
    qty: int,
    master_skus: set[str],
    prefix_map: dict[str, str],
    ambiguous_map: dict[str, list[str]],
    retail_lookup: dict[str, dict] | None = None,
):
    """
    Khớp một candidate SKU với master/retail data và cộng dồn vào counts.
    Logic khớp giống hệt TikTok extractor.
    """
    if candidate in master_skus:
        counts[candidate] += qty
    elif candidate.endswith('-') and candidate in prefix_map:
        counts[prefix_map[candidate]] += qty
    elif candidate.endswith('-') and candidate in ambiguous_map:
        # Ambiguous: chọn option đầu tiên làm fallback (không có context)
        # Với Shopee, các ambiguous prefix thường không xuất hiện
        pass
    elif retail_lookup and candidate in retail_lookup:
        counts[candidate] += qty
    else:
        print(f"   ⚠ SKU lạ: {candidate} (x{qty}) - không có trong combo lẫn retail")


# ============================================================
# PDF EXTRACTION — TIKTOK (Picking List)
# ============================================================

def _parse_pdf_header(full_text: str) -> dict:
    """
    Trich xuat thong tin header tu Picking List PDF.
    VD: Order quantity: 3 Product quantity: 3 Item quantity: 3
    Tra ve: {order_qty, product_qty, item_qty, print_time}
    """
    info = {}
    m = re.search(r'Order quantity:\s*(\d+)', full_text)
    if m:
        info['order_qty'] = int(m.group(1))
    m = re.search(r'Product quantity:\s*(\d+)', full_text)
    if m:
        info['product_qty'] = int(m.group(1))
    m = re.search(r'Item quantity:\s*(\d+)', full_text)
    if m:
        info['item_qty'] = int(m.group(1))
    m = re.search(r'Print time:\s*(.+)', full_text)
    if m:
        info['print_time'] = m.group(1).strip()
    return info


def extract_order_counts(
    pdf_path: str,
    master_skus: set[str],
    prefix_map: dict[str, str],
    ambiguous_map: dict[str, list[str]],
    retail_lookup: dict[str, dict] | None = None,
) -> tuple[dict[str, int], dict]:
    """
    Trich xuat so don hang cho moi Seller SKU tu PDF.
    Dung regex tim pattern: SellerSKU + Qty + OrderID (15+ chu so).

    Thu tu uu tien:
      1. Exact match trong master_skus (combo)
      2. Prefix match unique
      3. Prefix ambiguous -> doan tu context
      4. Tim trong retail_lookup (san pham don le)
      5. Khong tim thay -> bo qua + canh bao

    Tra ve: ({seller_sku: tong_so_qty}, {order_qty, product_qty, item_qty, print_time})
    """
    with pdfplumber.open(pdf_path) as pdf:
        texts = []
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                texts.append(t)
    full_text = "\n".join(texts)

    # Gộp các dòng để xử lý SKU ngắt dòng (thay \n = space)
    flat_text = full_text.replace('\n', ' ')

    # Pattern: Mã SKU (chứa ít nhất 1 chữ cái, có thể bắt đầu bằng số, có thể kết thúc bằng - nếu bị ngắt)
    #           + Qty (số đơn hàng) + OrderID (15+ chữ số)
    pattern = r'\b((?=[A-Za-z0-9]*[A-Za-z])[A-Za-z0-9]+(?:-[A-Za-z0-9]+)*(?:-)?)\s+(\d+)\s+(\d{15,})'

    counts: dict[str, int] = defaultdict(int)

    for match in re.finditer(pattern, flat_text):
        candidate = match.group(1)
        qty = int(match.group(2))

        if candidate in master_skus:
            # Exact match: Seller SKU xuat hien day du trong PDF
            counts[candidate] += qty
        elif candidate.endswith('-') and candidate in prefix_map:
            # Partial match unique: SKU bi ngat dong, tra prefix_map
            full_sku = prefix_map[candidate]
            counts[full_sku] += qty
        elif candidate.endswith('-') and candidate in ambiguous_map:
            # Prefix ambiguous: chon suffix DAI NHAT found trong context
            # (tranh bug: suffix "2" thang "BOG17-2" vi xuat hien truoc)
            options = ambiguous_map[candidate]
            after_text = flat_text[match.end():match.end()+200]
            best_match = None
            best_len = 0
            for opt in options:
                suffix = opt[len(candidate):]
                if suffix and suffix in after_text and len(suffix) > best_len:
                    best_match = opt
                    best_len = len(suffix)
            if best_match:
                counts[best_match] += qty
        elif candidate.endswith('-'):
            # Prefix khong co trong map nao -> bo qua
            pass
        elif retail_lookup and candidate in retail_lookup:
            # Tim thay trong san pham ban le
            counts[candidate] += qty
        else:
            # SKU la - khong co trong combo lan retail
            print(f"   ⚠ SKU la: {candidate} (x{qty}) - khong co trong ca 2 file")

    header_info = _parse_pdf_header(full_text)
    return dict(counts), header_info


# ============================================================
# CALCULATION
# ============================================================

def calculate_results(
    master_data: list[dict],
    order_counts: dict[str, int],
    retail_lookup: dict[str, dict] | None = None,
) -> list[dict]:
    """
    Doi chieu master_data voi order_counts tu PDF.
    - Seller SKU co trong master_data -> combo (theo dinh luong)
    - Seller SKU co trong retail_lookup -> san pham ban le
    - Khong tim thay -> canh bao + bo qua
    - Seller SKU không có trong master_data → sản phẩm đơn lẻ (1 đơn = 1 sp)
    """
    results = []
    matched_skus = set()
    for row in master_data:
        seller_sku = row["seller_sku"]
        if seller_sku in order_counts:
            matched_skus.add(seller_sku)
            mult = order_counts[seller_sku]
            # Tra cứu tên sản phẩm từ file retail (nếu SKU có trong đó)
            product_name = ''
            if retail_lookup:
                ri = retail_lookup.get(row["sku"])
                if ri:
                    product_name = ri.get("product_name", '')
            results.append({
                "seller_sku": seller_sku,
                "sku": row["sku"],
                "qty": row["qty"] * mult,
                "qty_sold": row["qty_sold"] * mult,
                "promo_qty": row["promo_qty"] * mult,
                "unit": row["unit"],
                "product_name": product_name,
            })

    # Sản phẩm bán lẻ: có trong PDF nhưng không có trong master_data combo
    for seller_sku, count in order_counts.items():
        if seller_sku not in matched_skus and retail_lookup and seller_sku in retail_lookup:
            r = retail_lookup[seller_sku]
            results.append({
                "seller_sku": seller_sku,
                "sku": r["sku"],
                "qty": r["qty"] * count,
                "qty_sold": r["qty_sold"] * count,
                "promo_qty": r["promo_qty"] * count,
                "unit": r["unit"],
                "product_name": r.get("product_name", ''),
            })
            matched_skus.add(seller_sku)

    # Cảnh báo SKU không tìm thấy ở đâu
    for seller_sku in order_counts:
        if seller_sku not in matched_skus:
            print(f"   ⚠ SKU không xác định: {seller_sku} (x{order_counts[seller_sku]}) - bỏ qua")

    return results



def generate_grouped_excel(results: list[dict], output_path: str, carrier: str = '', source_label: str = '', order_count: int = 0) -> str:
    """
    Tạo file Excel gộp theo SKU (không hiện Seller SKU).
    Format: SKU | Đơn vị tính | Qty | Qty Sold | Promo Qty
    Có dòng tiêu đề in đậm ở đầu để nhận diện khi in giấy.
    """
    # Gộp theo SKU
    grouped = {}
    for r in results:
        sku = r["sku"]
        if sku not in grouped:
            grouped[sku] = {"qty": 0, "qty_sold": 0, "promo_qty": 0, "unit": r.get("unit", ""), "product_name": r.get("product_name", "")}
        grouped[sku]["qty"] += r["qty"]
        grouped[sku]["qty_sold"] += r["qty_sold"]
        grouped[sku]["promo_qty"] += r["promo_qty"]
        # Giữ product_name đầu tiên khác rỗng
        if not grouped[sku]["product_name"] and r.get("product_name", ""):
            grouped[sku]["product_name"] = r["product_name"]

    grouped_list = [{"sku": k, **v} for k, v in grouped.items()]
    grouped_list.sort(key=lambda x: x["sku"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # ── Styles ──
    title_font = Font(name="Arial", size=16, bold=True, color="1F4E79")
    title_align = Alignment(horizontal="center", vertical="center")
    hdr_font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    data_font = Font(name="Arial", size=12)
    total_font = Font(name="Arial", size=12, bold=True)

    # ── Title row (dòng nhận diện khi in giấy) ──
    ncols = 7  # STT, SKU, Tên SP, ĐVT, SL, SL bán, SL KM
    if carrier and source_label:
        title_text = f'{carrier} — {source_label}'
    else:
        title_text = carrier or source_label or 'Báo cáo gộp SKU'
    if order_count > 0:
        title_text += f' — ({order_count} đơn)'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title_text)
    c.font = title_font
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28  # đủ cao để hiển thị rõ

    # ── Column headers (row 2) ──
    headers = ["STT", "SKU", "Tên sản phẩm", "Đơn vị tính", "SL", "SL bán", "SL KM"]
    # Độ rộng cột tối ưu: SKU=12, Tên SP=36 wrap, cột số=7 → tổng ~86 vừa A4 ngang
    col_widths = [5, 12, 36, 12, 7, 7, 7]
    col_aligns = ['C', 'C', 'L', 'C', 'R', 'R', 'R']  # center / left / right

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border

    # ── Data (bắt đầu từ row 3) ──
    for ri, r in enumerate(grouped_list):
        rn = ri + 3
        vals = [ri + 1, r["sku"], r.get("product_name", ""), r.get("unit", ""), r["qty"], r["qty_sold"], r["promo_qty"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=rn, column=ci, value=v)
            c.font = data_font
            c.border = thin_border
            # Căn lề theo cột + wrap text cho cột Tên sản phẩm (cột 3)
            ha = col_aligns[ci - 1] if ci <= len(col_aligns) else 'L'
            c.alignment = Alignment(horizontal={'C':'center','L':'left','R':'right'}.get(ha, 'left'),
                                    vertical='center',
                                    wrap_text=(ci == 3))

    # ── Total row ──
    tr = len(grouped_list) + 3
    tong_qty = sum(r["qty"] for r in grouped_list)
    tong_sold = sum(r["qty_sold"] for r in grouped_list)
    tong_promo = sum(r["promo_qty"] for r in grouped_list)

    ws.cell(row=tr, column=1, value="Tổng").font = total_font
    ws.cell(row=tr, column=1).border = thin_border
    ws.cell(row=tr, column=2).border = thin_border  # SKU trống
    ws.cell(row=tr, column=3).border = thin_border  # Tên sản phẩm trống
    ws.cell(row=tr, column=4).border = thin_border  # Đơn vị tính trống

    for ci, val in [(5, tong_qty), (6, tong_sold), (7, tong_promo)]:
        c = ws.cell(row=tr, column=ci, value=val)
        c.font = total_font
        c.border = thin_border

    # ── Column widths ──
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Print setup: vừa trang in, tránh mất cột ──
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # tự động số dòng
    ws.page_setup.paperSize = 9  # A4

    wb.save(output_path)
    print(f"   📊 Đã tạo Excel gộp: {os.path.basename(output_path)}")
    return output_path


# ============================================================
# EXPORT XLSX → PDF (via Excel COM)
# ============================================================

def export_xlsx_to_pdf(xlsx_path: str, pdf_path: str = '') -> str:
    """
    Dùng Excel COM để export file .xlsx sang .pdf.
    Yêu cầu: máy phải cài Microsoft Excel.
    Trả về đường dẫn file PDF, hoặc '' nếu thất bại.
    """
    import pythoncom
    import win32com.client
    import os as _os

    if not pdf_path:
        pdf_path = xlsx_path.rsplit('.', 1)[0] + '.pdf'

    excel = None
    workbook = None
    try:
        pythoncom.CoInitialize()
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        abs_xlsx = _os.path.abspath(xlsx_path)
        abs_pdf = _os.path.abspath(pdf_path)
        workbook = excel.Workbooks.Open(abs_xlsx)
        # 0 = xlTypePDF
        workbook.ExportAsFixedFormat(0, abs_pdf)
        print(f"   📄 Đã xuất PDF: {_os.path.basename(pdf_path)}")
        return pdf_path
    except Exception as e:
        print(f"   ⚠ Không thể xuất PDF từ Excel: {e}")
        return ''
    finally:
        try:
            if workbook is not None:
                workbook.Close(False)
        except Exception:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


# ============================================================
# REPORT AGGREGATION (tổng hợp nhiều file báo cáo)
# ============================================================

def extract_report_data(file_path: str) -> list[dict]:
    """
    Đọc một file báo cáo Phieu_xuat_hang_*.xlsx (đã được fill_template)
    và trích xuất dữ liệu SKU từ cả 2 panel (trái + phải).

    Trả về list[dict] dạng:
      {seller_sku, sku, qty, qty_sold, promo_qty, unit, product_name}
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    results: list[dict] = []

    for row_idx in range(3, ws.max_row + 1):
        # ── Panel trái: B=SKU, C=Tên SP, D=ĐVT, E=SL, F=SL bán, G=SL KM ──
        sku_left = ws.cell(row=row_idx, column=2).value  # Cột B
        if sku_left and str(sku_left).strip():
            sku = str(sku_left).strip().upper()
            product_name = str(ws.cell(row=row_idx, column=3).value or '').strip()
            unit = str(ws.cell(row=row_idx, column=4).value or '').strip()
            qty = _safe_int(ws.cell(row=row_idx, column=5).value)
            qty_sold = _safe_int(ws.cell(row=row_idx, column=6).value)
            promo_qty = _safe_int(ws.cell(row=row_idx, column=7).value)
            results.append({
                "seller_sku": sku,
                "sku": sku,
                "qty": qty,
                "qty_sold": qty_sold,
                "promo_qty": promo_qty,
                "unit": unit,
                "product_name": product_name,
            })

        # ── Panel phải: I=SKU, J=Tên SP, K=ĐVT, L=SL, M=SL bán, N=SL KM ──
        sku_right = ws.cell(row=row_idx, column=9).value  # Cột I
        if sku_right and str(sku_right).strip():
            sku = str(sku_right).strip().upper()
            product_name = str(ws.cell(row=row_idx, column=10).value or '').strip()
            unit = str(ws.cell(row=row_idx, column=11).value or '').strip()
            qty = _safe_int(ws.cell(row=row_idx, column=12).value)
            qty_sold = _safe_int(ws.cell(row=row_idx, column=13).value)
            promo_qty = _safe_int(ws.cell(row=row_idx, column=14).value)
            results.append({
                "seller_sku": sku,
                "sku": sku,
                "qty": qty,
                "qty_sold": qty_sold,
                "promo_qty": promo_qty,
                "unit": unit,
                "product_name": product_name,
            })

    wb.close()
    print(f"   📥 Đã trích xuất {len(results)} dòng từ: {os.path.basename(file_path)}")
    return results


def _safe_int(val) -> int:
    """Chuyển value sang int an toàn (None / str / float → int)."""
    if val is None:
        return 0
    try:
        return int(val)
    except (ValueError, TypeError):
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return 0


def aggregate_reports(
    report_files: list[str],
    output_dir: str,
    template_path: str,
) -> dict:
    """
    Tổng hợp nhiều file báo cáo Phieu_xuat_hang_*.xlsx thành 1 file duy nhất.
    - Trích xuất dữ liệu từ từng file
    - Gộp theo SKU, cộng dồn số lượng
    - Điền vào template và lưu ra file mới

    Trả về dict kết quả (cùng format với process_all).
    """
    from collections import defaultdict

    if not report_files:
        raise ValueError("Không có file báo cáo nào để tổng hợp")

    if not os.path.exists(template_path):
        raise ValueError(f"Không tìm thấy file template tại: {template_path}")

    # ── Trích xuất + gộp dữ liệu từ tất cả file ──
    merged: dict[str, dict] = {}  # {sku: {qty, qty_sold, promo_qty, product_name, unit}}
    total_files = 0
    total_order_count = 0

    for fp in report_files:
        if not os.path.exists(fp):
            print(f"   ⚠ Bỏ qua file không tồn tại: {fp}")
            continue
        try:
            rows = extract_report_data(fp)
            if not rows:
                print(f"   ⚠ Không trích xuất được dữ liệu từ: {os.path.basename(fp)}")
                continue
            total_files += 1
            # ── Trích xuất số đơn từ tiêu đề (hàng 1) ──
            try:
                from openpyxl import load_workbook
                wb_tmp = load_workbook(fp, data_only=True)
                ws_tmp = wb_tmp.active
                title_text = str(ws_tmp.cell(row=1, column=1).value or '')
                wb_tmp.close()
                m = re.search(r'SL đơn:\s*(\d+)', title_text)
                if m:
                    total_order_count += int(m.group(1))
            except Exception:
                pass
            for r in rows:
                sku = r["sku"]
                if sku not in merged:
                    merged[sku] = {
                        "qty": 0, "qty_sold": 0, "promo_qty": 0,
                        "product_name": "", "unit": "",
                    }
                merged[sku]["qty"] += r["qty"]
                merged[sku]["qty_sold"] += r["qty_sold"]
                merged[sku]["promo_qty"] += r["promo_qty"]
                if not merged[sku]["product_name"] and r.get("product_name", ""):
                    merged[sku]["product_name"] = r["product_name"]
                if not merged[sku]["unit"] and r.get("unit", ""):
                    merged[sku]["unit"] = r["unit"]
        except Exception as e:
            print(f"   ⚠ Lỗi xử lý {os.path.basename(fp)}: {e}")

    if not merged:
        raise ValueError("Không trích xuất được dữ liệu từ file nào")

    # ── Chuyển về định dạng results cho fill_template ──
    results = [
        {
            "seller_sku": sku,
            "sku": sku,
            "qty": info["qty"],
            "qty_sold": info["qty_sold"],
            "promo_qty": info["promo_qty"],
            "unit": info["unit"],
            "product_name": info["product_name"],
        }
        for sku, info in merged.items()
    ]

    print(f"\n📊 TỔNG HỢP {total_files} file, {total_order_count} đơn: {len(results)} SKU | "
          f"Qty={sum(r['qty'] for r in results)} | "
          f"Sold={sum(r['qty_sold'] for r in results)} | "
          f"Promo={sum(r['promo_qty'] for r in results)}")

    # ── Điền vào template ──
    now = datetime.now()
    output_path = os.path.join(output_dir, f"Phieu_xuat_hang_Tong_hop_{now.strftime('%m-%d_%H-%M-%S')}.xlsx")

    fill_template(results, template_path, output_path, carrier='Tổng hợp', order_count=total_order_count, source_label='Shopee')

    # ── Xuất PDF ──
    pdf_path = ''
    try:
        pdf_path = export_xlsx_to_pdf(output_path)
    except Exception as e:
        print(f"   ⚠ Không xuất được PDF: {e}")

    files_dict = {"xlsx_report": output_path}
    if pdf_path:
        files_dict["pdf_report"] = pdf_path

    return {
        "base_name": f"Tổng hợp ({total_files} file)",
        "rows": len(results),
        "tong_qty": sum(r["qty"] for r in results),
        "tong_sold": sum(r["qty_sold"] for r in results),
        "tong_promo": sum(r["promo_qty"] for r in results),
        "files": files_dict,
    }


# ============================================================
# TEMPLATE FILLING (điền số lượng vào mẫu Bảng thống kê hàng.xlsx)
# ============================================================

def fill_template(
    results: list[dict],
    template_path: str,
    output_path: str,
    carrier: str = '',
    order_count: int = 0,
    source_label: str = '',
) -> str:
    """
    Mở file mẫu Bảng thống kê hàng.xlsx, điền số lượng (SL, SL bán, SL KM)
    vào các dòng có SKU khớp, rồi lưu ra file mới.
    Các SKU không có sẵn trong template sẽ được tự động thêm vào dòng trống
    hoặc append vào cuối.

    Template có 2 panel:
      - Trái:  cột B = Mã sản phẩm,  E = SL,  F = SL bán,  G = SL KM
      - Phải:  cột I = Mã sản phẩm,  L = SL,  M = SL bán,  N = SL KM
    """
    from openpyxl import load_workbook
    from copy import copy

    wb = load_workbook(template_path)
    ws = wb.active

    # ── Build lookup từ results: key = sku (và cả seller_sku) ──
    # Gộp theo sku (phòng trường hợp nhiều dòng cùng sku)
    grouped: dict[str, dict] = {}
    for r in results:
        sku = r["sku"]
        if sku not in grouped:
            grouped[sku] = {"qty": 0, "qty_sold": 0, "promo_qty": 0, "product_name": "", "unit": ""}
        grouped[sku]["qty"] += r["qty"]
        grouped[sku]["qty_sold"] += r["qty_sold"]
        grouped[sku]["promo_qty"] += r["promo_qty"]
        if not grouped[sku]["product_name"] and r.get("product_name", ""):
            grouped[sku]["product_name"] = r["product_name"]
        if not grouped[sku]["unit"] and r.get("unit", ""):
            grouped[sku]["unit"] = r["unit"]

    # Build index: ánh xạ seller_sku → sku (để match cả combo code)
    seller_to_sku: dict[str, str] = {}
    for r in results:
        if r["seller_sku"] != r["sku"]:
            seller_to_sku[r["seller_sku"]] = r["sku"]

    def find_qty(sku_code: str):
        """Tìm số lượng cho một mã SKU bất kỳ (có thể là seller_sku hoặc sku)."""
        if sku_code in grouped:
            return grouped[sku_code]
        if sku_code in seller_to_sku:
            real_sku = seller_to_sku[sku_code]
            if real_sku in grouped:
                return grouped[real_sku]
        return None

    # ── Panel trái: cột B (SKU), cột E (SL), F (SL bán), G (SL KM) ──
    matched_skus: set[str] = set()
    filled_left = 0
    for row_idx in range(3, ws.max_row + 1):
        sku_cell = ws.cell(row=row_idx, column=2)  # Cột B
        sku_code = str(sku_cell.value).strip() if sku_cell.value else ''
        if not sku_code:
            continue
        qty_info = find_qty(sku_code)
        if qty_info:
            ws.cell(row=row_idx, column=5, value=qty_info["qty"])       # SL
            ws.cell(row=row_idx, column=6, value=qty_info["qty_sold"])  # SL bán
            ws.cell(row=row_idx, column=7, value=qty_info["promo_qty"]) # SL KM
            filled_left += 1
            matched_skus.add(sku_code)

    # ── Panel phải: cột I (SKU), cột L (SL), M (SL bán), N (SL KM) ──
    filled_right = 0
    for row_idx in range(3, ws.max_row + 1):
        sku_cell = ws.cell(row=row_idx, column=9)  # Cột I
        sku_code = str(sku_cell.value).strip() if sku_cell.value else ''
        if not sku_code:
            continue
        qty_info = find_qty(sku_code)
        if qty_info:
            ws.cell(row=row_idx, column=12, value=qty_info["qty"])       # SL
            ws.cell(row=row_idx, column=13, value=qty_info["qty_sold"])  # SL bán
            ws.cell(row=row_idx, column=14, value=qty_info["promo_qty"]) # SL KM
            filled_right += 1
            matched_skus.add(sku_code)

    # ── Thêm các SKU chưa có trong template vào dòng trống hoặc append cuối ──
    unmatched: list[dict] = []
    for sku, info in grouped.items():
        # Kiểm tra cả sku và seller_sku đều chưa được match
        if sku not in matched_skus:
            # Tìm seller_sku tương ứng
            seller = sku
            for s, r in seller_to_sku.items():
                if r == sku:
                    seller = s
                    break
            if seller not in matched_skus:
                unmatched.append({"sku": sku, **info})

    added_count = 0
    if unmatched:
        # Lấy style từ dòng dữ liệu cuối cùng có SKU để áp dụng cho dòng mới
        # Tìm dòng tham chiếu (dòng 3)
        ref_row = 3

        # Xác định max_row hiện tại
        current_max = ws.max_row

        for item in unmatched:
            # Tìm vị trí trống đầu tiên trong panel trái hoặc phải
            placed = False

            # Ưu tiên điền vào dòng trống có sẵn (có STT nhưng chưa có SKU)
            for row_idx in range(3, current_max + 1):
                # Panel trái: ô B trống
                b_val = ws.cell(row=row_idx, column=2).value
                if not b_val or str(b_val).strip() == '':
                    # Điền vào panel trái
                    ws.cell(row=row_idx, column=2, value=item["sku"])
                    ws.cell(row=row_idx, column=3, value=item.get("product_name", ""))
                    ws.cell(row=row_idx, column=4, value=item.get("unit", ""))
                    ws.cell(row=row_idx, column=5, value=item["qty"])
                    ws.cell(row=row_idx, column=6, value=item["qty_sold"])
                    ws.cell(row=row_idx, column=7, value=item["promo_qty"])
                    placed = True
                    added_count += 1
                    break

                # Panel phải: ô I trống
                i_val = ws.cell(row=row_idx, column=9).value
                if not i_val or str(i_val).strip() == '':
                    # Điền vào panel phải
                    ws.cell(row=row_idx, column=9, value=item["sku"])
                    ws.cell(row=row_idx, column=10, value=item.get("product_name", ""))
                    ws.cell(row=row_idx, column=11, value=item.get("unit", ""))
                    ws.cell(row=row_idx, column=12, value=item["qty"])
                    ws.cell(row=row_idx, column=13, value=item["qty_sold"])
                    ws.cell(row=row_idx, column=14, value=item["promo_qty"])
                    placed = True
                    added_count += 1
                    break

            # Nếu không còn dòng trống → append dòng mới vào cuối
            if not placed:
                new_row = current_max + 1
                ws.cell(row=new_row, column=1, value=new_row - 2)  # STT
                ws.cell(row=new_row, column=2, value=item["sku"])
                ws.cell(row=new_row, column=3, value=item.get("product_name", ""))
                ws.cell(row=new_row, column=4, value=item.get("unit", ""))
                ws.cell(row=new_row, column=5, value=item["qty"])
                ws.cell(row=new_row, column=6, value=item["qty_sold"])
                ws.cell(row=new_row, column=7, value=item["promo_qty"])
                current_max = new_row
                added_count += 1

    # ── Cập nhật tiêu đề (dòng 1) ──
    now = datetime.now()
    carrier_str = carrier if carrier else ''
    title_parts = ['Phiếu xuất hàng ngày :']
    title_parts.append(f'SL đơn: {order_count}')
    if carrier_str:
        title_parts.append(f'ĐVVC: {carrier_str}')
    if source_label:
        title_parts.append(source_label)
    title_parts.append(now.strftime('%d/%m/%Y %H:%M'))
    ws.cell(row=1, column=1).value = '    '.join(title_parts)

    # ── Cập nhật print area & page setup để in cả 2 cột ──
    # Template gốc có print area chỉ $B$1:$G$29 (cột trái), cần mở rộng ra cả cột phải
    last_col_letter = 'N'
    # Set print area với sheet name để ghi đè triệt để template cũ
    ws.print_area = f"'{ws.title}'!$A$1:${last_col_letter}${ws.max_row}"
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = 9  # A4

    # ── Lề giấy hẹp để tận dụng tối đa không gian in ──
    ws.page_margins.left = 0.25     # ~0.6 cm
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.3       # ~0.8 cm
    ws.page_margins.bottom = 0.3
    ws.page_margins.header = 0.0
    ws.page_margins.footer = 0.0

    # ── Set độ rộng cột để lấp đầy trang A4 ngang ──
    # Tổng ~162 — vừa vùng in A4 ngang với margin 0.25"
    col_widths = {
        'A': 4,  'B': 11, 'C': 24, 'D': 10, 'E': 9, 'F': 9, 'G': 9,
        'H': 4,  'I': 11, 'J': 22, 'K': 10, 'L': 9, 'M': 9, 'N': 9,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ── Set độ cao hàng ──
    ws.row_dimensions[1].height = 22   # Tiêu đề
    ws.row_dimensions[2].height = 26   # Header cột
    for r in range(3, ws.max_row + 1):
        ws.row_dimensions[r].height = 18   # Dòng dữ liệu

    # ── Áp dụng font + alignment cho header (row 2) ──
    hdr_font = Font(name='Arial', size=12, bold=True)
    for col_idx in range(1, 15):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # ── Áp dụng font + alignment cho TOÀN BỘ data cells ──
    data_font = Font(name='Arial', size=10)
    data_align_center = Alignment(horizontal='center', vertical='center')
    data_align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    # Cột tên sản phẩm (C & J) cần wrap text + left align
    wrap_cols = {3, 10}
    for row_idx in range(3, ws.max_row + 1):
        for col_idx in range(1, 15):  # A=1 đến N=14
            cell = ws.cell(row=row_idx, column=col_idx)
            # Chỉ style nếu ô có dữ liệu hoặc là ô trống trong vùng in
            cell.font = data_font
            cell.border = thin_border
            if col_idx in wrap_cols:
                cell.alignment = data_align_left
            else:
                cell.alignment = data_align_center

    wb.save(output_path)
    wb.close()

    total_filled = filled_left + filled_right
    print(f'   📊 Đã điền số lượng cho {total_filled} SKU (trái={filled_left}, phải={filled_right}) vào template')
    if added_count > 0:
        print(f'   ➕ Đã thêm {added_count} SKU mới không có sẵn trong template')
    print(f'   💾 Đã lưu: {os.path.basename(output_path)}')
    return output_path


# ============================================================
# MAIN ENTRY POINT
# ============================================================

def process_all(
    pdf_files: list[str],
    output_dir: str,
    master_path: str | None = None,
    retail_path: str | None = None,
    carrier: str = '',
    template_path: str | None = None,
) -> list[dict]:
    """
    Xu ly toan bo pipeline cho nhieu file PDF cung 1 carrier.
    TAT CA PDF duoc gop chung vao 1 file bao cao duy nhat.
    Ket qua duoc dien vao template Bảng thống kê hàng.xlsx.
    """
    if master_path is None:
        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        master_path = os.path.join(base, "master_data.xlsx")

    if not os.path.exists(master_path):
        raise ValueError(f"Khong tim thay file master_data tai: {master_path}")

    # Tự động tìm template nếu không được chỉ định
    if template_path is None:
        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        template_path = os.path.join(base, "Bảng thống kê hàng.xlsx")

    if not os.path.exists(template_path):
        raise ValueError(f"Khong tim thay file template tai: {template_path}")

    # Load master data (combo)
    print(f"Load master data: {master_path}")
    master_data = load_master_data(master_path)
    master_skus, prefix_map, ambiguous_map = build_sku_index(master_data)
    print(f"   {len(master_data)} dong, {len(master_skus)} Seller SKU"
          + (f", {len(prefix_map)} prefix map" if prefix_map else ""))

    # Load retail data
    retail_lookup = None
    if retail_path and os.path.exists(retail_path):
        retail_lookup = load_retail_data(retail_path)
    elif retail_path:
        print(f"   ⚠ Khong tim thay file retail: {retail_path}")

    # Gom order_counts tu TAT CA PDF
    from collections import defaultdict
    merged_order_counts: dict[str, int] = defaultdict(int)
    total_order_qty = 0   # Số đơn hàng thực tế (Order quantity từ header)
    all_order_sns: set[str] = set()  # Gom tất cả Order SN

    for pdf_path in pdf_files:
        print(f"\nXu ly: {os.path.basename(pdf_path)}")

        # ── Auto-detect loại PDF ──
        pdf_type = detect_pdf_type(pdf_path)
        print(f"   Loai PDF: {pdf_type.upper()}")

        if pdf_type == 'shopee':
            # ── Tự động tách phiếu xuất & shipping label nếu là file gộp ──
            picking_path, shipping_path, shipping_pages = split_shopee_pdf(pdf_path, output_dir)
            if picking_path:
                # Dùng file phiếu xuất đã tách để tính toán
                calc_path = picking_path
                print(f"   → Dùng phiếu xuất để tính: {os.path.basename(picking_path)}")
            else:
                calc_path = pdf_path
                shipping_pages = 0

            if shipping_path:
                print(f"   → File shipping label: {os.path.basename(shipping_path)}")

            order_counts, header_info = extract_order_counts_shopee(
                calc_path, master_skus, prefix_map, ambiguous_map, retail_lookup)
            # SL đơn = số shipping label (mỗi label = 1 đơn)
            order_qty = shipping_pages if shipping_pages > 0 else header_info.get('order_qty', 0)
        else:
            # TikTok hoặc unknown → dùng extractor TikTok (regex-based)
            order_counts, header_info = extract_order_counts(
                pdf_path, master_skus, prefix_map, ambiguous_map, retail_lookup)

        if not order_counts:
            print(f"   ⚠ Khong tim thay Seller SKU nao trong PDF!")
            continue
        for sku, count in order_counts.items():
            merged_order_counts[sku] += count
        if pdf_type != 'shopee':
            order_qty = header_info.get('order_qty', 0)
        total_order_qty += order_qty
        # Gom Order SN ĐẦY ĐỦ từ shipping label (dài hơn picking list)
        if pdf_type == 'shopee' and shipping_path and os.path.exists(shipping_path):
            full_sns = _extract_order_sns_from_shipping(shipping_path)
            all_order_sns.update(full_sns)
        elif header_info.get('order_sns'):
            # Fallback: lấy từ picking list (có thể thiếu vài ký tự cuối)
            all_order_sns.update(header_info['order_sns'])
        print(f"   Tim thay {len(order_counts)} Seller SKU, {order_qty} don hang (header), {sum(order_counts.values())} mat hang")

    if not merged_order_counts:
        raise ValueError("Không trích xuất được dữ liệu từ bất kỳ PDF nào. "
                         "Kiểm tra file đầu vào và master_data.")

    # Tinh toan tu merged order counts
    results = calculate_results(master_data, dict(merged_order_counts), retail_lookup)
    print(f"\n📊 KET QUA CHUNG: {len(results)} dong | "
          f"Qty={sum(r['qty'] for r in results)} | "
          f"Sold={sum(r['qty_sold'] for r in results)} | "
          f"Promo={sum(r['promo_qty'] for r in results)}")

    # Điền số lượng vào template thay vì tạo PDF
    now = datetime.now()
    carrier_safe = carrier.replace(' ', '_').replace('&', 'n') if carrier else ''
    prefix = f"Phieu_xuat_hang_{carrier_safe}_" if carrier_safe else "Phieu_xuat_hang_"
    output_path = os.path.join(output_dir, f"{prefix}{now.strftime('%m-%d_%H-%M-%S')}.xlsx")

    fill_template(results, template_path, output_path, carrier=carrier, order_count=total_order_qty, source_label='Shopee')

    # ── Lưu danh sách Order SN ra file .txt ──
    if all_order_sns:
        order_sn_path = os.path.join(output_dir, f"{prefix}Order_SN_{now.strftime('%m-%d_%H-%M-%S')}.txt")
        with open(order_sn_path, 'w', encoding='utf-8') as f:
            f.write(f"# Order SN — {carrier or 'tat ca'} — {now.strftime('%d/%m/%Y %H:%M:%S')}\n")
            f.write(f"# Tong: {len(all_order_sns)} Order SN\n")
            f.write(f"# SL don: {total_order_qty}\n")
            f.write(f"# SL mat hang: {sum(merged_order_counts.values())}\n\n")
            for sn in sorted(all_order_sns):
                f.write(sn + '\n')
        print(f"   📋 Đã lưu {len(all_order_sns)} Order SN → {os.path.basename(order_sn_path)}")

        # ── Gửi Order SN lên API ──
        try:
            import urllib.request
            sns_csv = ','.join(sorted(all_order_sns))
            data = sns_csv.encode('utf-8')
            req = urllib.request.Request(
                'http://88.2.0.55:7016/api/ids/receive',
                data=data,
                headers={'Content-Type': 'text/plain'},
                method='POST',
            )
            with urllib.request.urlopen(req, timeout=30) as resp:
                print(f"   📡 Đã gửi {len(all_order_sns)} Order SN lên API — HTTP {resp.status}")
        except Exception as e:
            print(f"   ⚠ Không gửi được Order SN lên API: {e}")

    # ── Xuất PDF từ file Excel vừa tạo ──
    pdf_path = ''
    try:
        pdf_path = export_xlsx_to_pdf(output_path)
    except Exception as e:
        print(f"   ⚠ Không xuất được PDF: {e}")

    files_dict = {"xlsx_report": output_path}
    if pdf_path:
        files_dict["pdf_report"] = pdf_path
    if all_order_sns:
        files_dict["order_sn_txt"] = order_sn_path

    # ── Dọn file tách trung gian (_phieu_xuat.pdf) ──
    cleanup_count = 0
    for pdf_path in pdf_files:
        stem = os.path.splitext(os.path.basename(pdf_path))[0]
        # Xóa file phiếu xuất đã tách (không cần giữ)
        picking_f = os.path.join(output_dir, f'{stem}_phieu_xuat.pdf')
        if os.path.exists(picking_f):
            try:
                os.remove(picking_f)
                cleanup_count += 1
            except Exception:
                pass
    if cleanup_count > 0:
        print(f'   🗑 Đã dọn {cleanup_count} file phiếu xuất trung gian')

    return [{
        "base_name": f"Combined {carrier or 'all'}",
        "rows": len(results),
        "tong_qty": sum(r["qty"] for r in results),
        "tong_sold": sum(r["qty_sold"] for r in results),
        "tong_promo": sum(r["promo_qty"] for r in results),
        "files": files_dict,
    }]


# ============================================================
# CLI — Test nhanh
# ============================================================
if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python calculator.py <pdf_file> [pdf_file2...]")
        sys.exit(1)

    pdfs = sys.argv[1:]
    out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "outputs")
    os.makedirs(out_dir, exist_ok=True)

    master = os.path.join(os.path.dirname(os.path.abspath(__file__)), "master_data.xlsx")

    try:
        results = process_all(pdfs, out_dir, master)
        print("\n" + "=" * 60)
        print("✅ HOÀN THÀNH!")
        for r in results:
            print(f"\n📦 {r['base_name']}:")
            print(f"   Dòng: {r['rows']} | Qty: {r['tong_qty']} | "
                  f"Sold: {r['tong_sold']} | Promo: {r['tong_promo']}")
            for k, v in r["files"].items():
                print(f"   📎 {k}: {os.path.basename(v)}")
    except ValueError as e:
        print(f"\n❌ Lỗi: {e}")
        sys.exit(1)
